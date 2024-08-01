import base64
from io import BytesIO
import re
import numpy as np
import openpyxl
import pandas as pd
import os
from PIL import Image
from html2image import Html2Image
import requests
from bs4 import BeautifulSoup
import threading
from datetime import datetime
import json
from concurrent.futures import ThreadPoolExecutor
import pyfiglet
import urllib3
from queue import Queue
import time
from rich import print, pretty
from rich.progress import track, Progress
from rich.table import Table
from rich.console import Console
import sys
import logging
import shutil
import calendar
import pyperclip
import fitz
from pypdf import PdfReader, PdfWriter

# # set cmd to support utf-8
# os.system('chcp 874')
# os.system('$LANG="th_TH.UTF-8"')
# os.system('cls')

# from closePM import closePM
# from closeCAL import closeCAL

from openpyxl import load_workbook
logging.basicConfig(filename='log.txt', filemode='a', format='%(asctime)s - %(levelname)s - %(message)s',
                    level=logging.ERROR)
# logging.basicConfig(level=logging.ERROR)


# pretty.install()
# pretty.pretty_print = True

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

init_text = pyfiglet.figlet_format("BME Assistant", font="slant")
print(init_text)
root_dir = os.path.dirname(os.path.abspath(__file__))

# check if using pyinstaller
if getattr(sys, 'frozen', False):
    root_dir = os.path.dirname(sys.executable)
config = open(os.path.join(root_dir, "CONFIG", "config.json"), "r")
confdata = json.load(config)

login_site = confdata["SITE"]

lock = threading.Lock()
data = {
    "user": confdata["USERNAME"],
    "pass": confdata["PASSWORD"],
    "Submit": "Submit",
    "Submit.x": "79",
    "Submit.y": "30",
}

cookies = {
    "_ga": "GA1.1.409909798.1624509466",
    "_ga_L9CPT990SV": "GS1.1.1629863162.2.0.1629863162.0",
    "PHPSESSID": confdata["SESSION_ID"],
}


headers = {
    "Connection": "keep-alive",
    "Upgrade-Insecure-Requests": "1",
    "User-Agent": "Mozilla/5.0 (Windows NT 6.1; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/99.0.4844.51 Safari/537.36",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9",
    "Referer": "http://nsmart.nhealth-asia.com/mtdpdb01/default.php",
    "Accept-Language": "th-TH,th;q=0.9,en;q=0.8",
}
emp_list = {}
tools_list = {}
calibrator_list = {}
temp_emp_list = {}
SEARCH_KEY = 's_code'

excel_file_name = 'recordCal_PM.xlsx'
excel_folder = os.path.join(root_dir, 'EXCEL FILE')


def read_excel_file():
    global master_df
    global SEARCH_KEY
    master_df = (pd.read_excel(
        os.path.join(excel_folder, excel_file_name).replace('//', '/'), sheet_name="Sheet1", engine='openpyxl')).dropna(subset=[SEARCH_KEY[0]])
    # check if dataframe gasattr map
    if not hasattr(master_df, 'map'):
        master_df.map = master_df.applymap
    return master_df.map(str)


def internet_connection():
    try:
        requests.get('https://www.google.com/', verify=False)
        return True
    except requests.exceptions.RequestException as e:
        return False


def set_login():
    global confdata
    global cookies
    try:
        with requests.Session() as s:
            r = s.post("https://nsmart.nhealth-asia.com/MTDPDB01/index.php",
                    data=data,
                    headers=headers,
                    verify=False)
            confdata["SESSION_ID"] = s.cookies['PHPSESSID']
            cookies['PHPSESSID'] = s.cookies['PHPSESSID']
            # write over file to clear old data
            with open(os.path.join(root_dir, "CONFIG/config.json"), "w") as json_file:
                json.dump(confdata, json_file)

    except Exception as e:
        print(f"An error occurred on line {sys.exc_info()[-1].tb_lineno}: {e}")
        print("[red]No Internet Connection[/red]")
        time.sleep(5)
        set_login()


def formatDate(date):
    if date != 'nan':
        # date is dd/mm/yyyy
        date = date.split('/')
        date[0] = '0' + date[0] if len(date[0]) == 1 else date[0]
        date[1] = '0' + date[1] if len(date[1]) == 1 else date[1]
        date = '/'.join(date)
    return date

def convertTime(time):
    if time != 'nan':
        # convert time is second to text format
        hours = time / 3600
        minutes = (time % 3600) / 60
        seconds = time % 60
        time = f'{int(hours)} hours {int(minutes)} minutes {int(seconds)} seconds'
        return time

    return time
def save_empList(emp_list=None):
    # write over file to clear old data
    if emp_list is None:
        return
    with open(os.path.join(root_dir, 'CONFIG/emp_list.json'), 'w') as json_file:
        json.dump(emp_list, json_file)


def save_calibrator_list(calibrator_list=None):
    # write over file to clear old data
    if calibrator_list is None:
        return
    with open(os.path.join(root_dir, 'CONFIG/calibrator_list.json'), 'w') as json_file:
        json.dump(calibrator_list, json_file)


def load_empList():
    global emp_list
    # handle if file not found
    try:
        with open(os.path.join(root_dir, 'CONFIG/emp_list.json'), 'r') as json_file:
            emp_list = json.load(json_file)
        if emp_list is None or len(emp_list) == 0:
            get_emp_list()
    except Exception as e:
        get_emp_list()
        



def load_tool_list(href):
    global tools_list
    # handle if file not found
    if href is not None and href != 'none':
        print('Load tools list')
        try:
            response = requests.get(
                "https://nsmart.nhealth-asia.com/MTDPDB01/pm/maintain07.php?" +
                href + "&ccsForm=main_jobs%3AEdit",
                headers=headers,
                cookies=cookies,
                verify=False,
            )
        except requests.exceptions.RequestException as e:
            print(
                f"An error occurred on line {sys.exc_info()[-1].tb_lineno}: {e}")
            time.sleep(5)
            return load_tool_list(href)
        response.encoding = "tis-620"
        soup = BeautifulSoup(response.text, "lxml")
        select = soup.find('select', {'name': 'toolid'})
        if len(select) == 0:
            print('No tools list found')
            exit()
        tools_list[confdata['SITE'].lower()] = {}
        options = select.findAll('option')
        for option in options:
            tools_list[confdata['SITE'].lower(
            )][option.text.strip().lower()] = option['value']
        save_tool_list(tools_list)
    try:
        with open(os.path.join(root_dir, 'CONFIG/tools_list.json'), 'r') as json_file:
            tools_list = json.load(json_file)
        if tools_list is None or len(tools_list) == 0:
            tools_list = {}
            save_tool_list(tools_list)
    except Exception as e:
        tools_list = {}
        save_tool_list(tools_list)


def save_tool_list(tools_list=None):
    # write over file to clear old data
    if tools_list is None:
        return
    with open(os.path.join(root_dir, 'CONFIG/tools_list.json'), 'w') as json_file:
        json.dump(tools_list, json_file)


def load_calibrator_list():
    global calibrator_list
    # handle if file not found
    try:
        with open(os.path.join(root_dir, 'CONFIG/calibrator_list.json'), 'r') as json_file:
            calibrator_list = json.load(json_file)
        if calibrator_list is None or len(calibrator_list) == 0:
            print('No calibrator list found')
            exit()
    except Exception as e:
        print(f"An error occurred on line {sys.exc_info()[-1].tb_lineno}: {e}")
        get_emp_list()

equipments_arr = []
def get_equipment_file(url='https://nsmart.nhealth-asia.com/MTDPDB01/asset_mast_list_new.php?asset_masterPageSize=100', page='1'):
    global equipments_arr
    if page == '1':
        equipments_arr = []
    page_url = url + "&asset_masterPage=" + str(page)
    try:
        response = requests.get(
            page_url,
            headers=headers,
            cookies=cookies,
            verify=False,
        )
    except requests.exceptions.RequestException as e:
        print(f"An error occurred on line {sys.exc_info()[-1].tb_lineno}: {e}")
        time.sleep(5)
        return get_equipment_file(url, page)
    response.encoding = "tis-620"
    # find table with regex    
    res = re.findall(r'<table\s+[^>]*class=["\']Grid["\'][^>]*>', response.text, re.DOTALL)
    if len(res) == 0:
        print("No table found, re-login...")
        set_login()
        return get_equipment_file(url, page)
    # print attr class of each table
    # find index of '<table class="Grid" cellspacing="0" cellpadding="0">'
    # index = response.text.find('<table class="Grid" cellspacing="0" cellpadding="0">')
    tmp = response.text.split('<table class="Grid" cellspacing="0" cellpadding="0">')
    tmp2 = tmp[1].split('</table>')
    table = '<table class="Grid" cellspacing="0" cellpadding="0">' + tmp2[0] + '</table>'
    soup = BeautifulSoup(table, "lxml")
    max_page = soup.find_all('tr', {'class': 'Footer'})[0].text.split('of')[1].strip().split(' ')[0]
    print('[yellow]Fetching page[/yellow] [blue]{}[/blue] [yellow]out of[/yellow] [blue]{}[/blue]' .format(page, max_page))
    rows = soup.find_all('tr')
    header = [ele.text.strip() for ele in rows[1].find_all('th')]
    if page == 1:
        rows = rows[1:-1]
    else:
        rows = rows[2:-1]
    for row in rows:
        cols = row.find_all('td')
        img = ('https://nsmart.nhealth-asia.com/MTDPDB01/' + cols[1].find('a').get('href')) if cols[1].find('a') is not None else ''
        cols = [ele.text.strip() if ele.text.strip() != 'Click' else '' for ele in cols]
        cols[1] = img
        equipments_arr.append(cols)
    if int(page) >= int(max_page):
    # if int(page) == 2:
        df = pd.DataFrame(equipments_arr, columns=header)
        df.to_excel(os.path.join(root_dir,'EXCEL FILE', 'equipment_list.xlsx'), index=False)
        # print file path to let user to click
        print('[yellow]The equipment list file has been saved at[/yellow] [blue]{}[/blue]' .format(os.path.join(root_dir, 'EXCEL FILE', 'equipment_list.xlsx')))
        # open file
        file_path = os.path.join(root_dir, 'EXCEL FILE', 'equipment_list.xlsx')
        auto_adjust_column_width_from_df(file_path, 'Sheet1')
        os.system('start excel.exe "' + file_path+ '"')
        
        return equipments_arr
    else:
        
        page = int(page) + 1
        return get_equipment_file(url, str(page))
    
def auto_adjust_column_width_from_df(file_path, sheet_name):
    # Load the workbook and select the specified sheet
    workbook = load_workbook(file_path)
    worksheet = workbook[sheet_name]

    for column in worksheet.columns:
        max_length = 0
        for cell in column:
            try:
                # Check the length of the cell value
                cell_length = len(str(cell.value))
                max_length = max(max_length, cell_length)
            except:
                pass
        # Adjust the column width; add a little extra space
        adjusted_width = (max_length + 2)
        worksheet.column_dimensions[openpyxl.utils.get_column_letter(column[0].column)].width = adjusted_width

    # Save the workbook
    workbook.save(file_path)


temp_team_list = {}
def get_team_list(url, page='1'):
    global temp_team_list
    if page == '1':
        temp_team_list = {}
    page_url = url + "&employeePage=" + str(page)
    try:
        response = requests.get(
            page_url,
            headers=headers,
            cookies=cookies,
            verify=False,
        )
    except requests.exceptions.RequestException as e:
        print(f"An error occurred on line {sys.exc_info()[-1].tb_lineno}: {e}")
        time.sleep(5)
        return get_team_list(url, page)
    response.encoding = "tis-620"
    soup = BeautifulSoup(response.text, "lxml")
    form = soup.find('form', {'name': 'employee'})
    if form is None:
        print("No form found, re-login...")
        set_login()
        return get_team_list(url, page)

    tr = form.find_all('tr', {'class': 'Row'})
    footer = form.find_all('tr', {'class': 'Footer'})[
        0].text.strip().split('of')[1].strip().split(' ')[0]
    if page > footer:
        return temp_team_list
    max_emp = int(tr[0].find('td').text.strip().split('\xa0')[1])
    if len(temp_team_list) >= max_emp:
        return temp_team_list
    tr.pop(0)
    if max_emp > 0:
        for row in tr:
            input = row.find_all('input')
            if input[0]['value'] == '' or input[1]['value'] == '':
                continue
            # temp_emp_list.append([input[0]['value'], input[1]['value']])
            temp_team_list[input[1]['value'].lower().replace(
                '  ', ' ')] = input[0]['value']
        # get next page
        page = int(page) + 1
        return get_team_list(url, str(page))
    else:
        return temp_team_list


def get_emp_list():
    global emp_list
    try:
        response = requests.get(
            "https://nsmart.nhealth-asia.com/MTDPDB01/reftable/employee_branch.php?dept_control=1",
            headers=headers,
            cookies=cookies,
            verify=False,
        )
    except requests.exceptions.RequestException as e:
        print(f"An error occurred on line {sys.exc_info()[-1].tb_lineno}: {e}")
        time.sleep(5)
        return get_emp_list()
    response.encoding = "tis-620"
    soup = BeautifulSoup(response.text, "lxml")
    form = soup.find('form', {'name': 'm_dept_tech'})
    if form is None:
        print("No form found, re-login...")
        set_login()
        return get_emp_list()

    tr = form.find_all('tr', {'class': 'Row'})
    tr.pop(0)
    for row in tr:
        td = row.find_all('td')
        calibrator_list[td[1].text.strip().lower()] = td[0].text.strip()
        print('[yellow]Fetching members of team[/yellow] [blue]{}[/blue]' .format(td[1].text.strip()))
        url = "https://nsmart.nhealth-asia.com/MTDPDB01/reftable/employee_branch.php?dept_control=1&dept_tech={}".format(
            td[0].text.strip())
        team_list = get_team_list(url)
        emp_list[td[1].text.strip().lower()] = team_list
        emp_list[td[1].text.strip().lower()]["option_name"] = td[1].text.strip()
    save_calibrator_list(calibrator_list)
    save_empList(emp_list)


def getFirstAndLastDay(date):
    if date is None or date == '':
        date = datetime.now()
    # check if date is string
    if isinstance(date, str):
        date = datetime.strptime(date, '%d/%m/%Y')
    # get first and last day of month
    first_day = date.replace(day=1)
    last_day = date.replace(day=calendar.monthrange(date.year, date.month)[1])
    first_day = first_day.strftime("%d/%m/%Y")
    last_day = last_day.strftime("%d/%m/%Y")
    date_year = first_day.split('/')[2]
    return first_day, last_day, date_year


def base64_image_to_base64_pdf(base64_image_string):
    try:
        # Decode the base64 image string to bytes
        image_data = base64.b64decode(base64_image_string)

        # Create a BytesIO object to open the image with Pillow
        image_buffer = BytesIO(image_data)

        # Open the image using Pillow
        image = Image.open(image_buffer)

        # Create a BytesIO object to save the PDF
        pdf_buffer = BytesIO()
        image.save(pdf_buffer, "PDF", resolution=100.0, save_all=True)

        # Get the content of the PDF as bytes
        pdf_bytes = pdf_buffer.getvalue()

        # Encode the PDF as a base64 string
        pdf_base64 = base64.b64encode(pdf_bytes).decode("utf-8")

        return pdf_base64

    except Exception as e:
        # print error with line number
        print(f"An error occurred on line {sys.exc_info()[-1].tb_lineno}: {e}")
        return None


def get_screen_shot(soup, css_file, text):
    # open css file from css folder
    css_file = open(os.path.join(root_dir, 'SOURCE/'+css_file), 'r')
    css = css_file.read()

    # get file path
    __location__ = os.path.join(root_dir, 'SCREENSHOT')
    # if folder not exist, create folder
    if not os.path.exists(__location__):
        os.makedirs(__location__)
    hti = Html2Image(temp_path=__location__, size=(
        800, 800), disable_logging=True)
    file_name = str(datetime.timestamp(datetime.now())*1000)
    try:
        hti.screenshot(html_str=str(soup), css_str=css,
                    save_as=file_name+'.png')
    except Exception as e:
        print(f"An error occurred on line {sys.exc_info()[-1].tb_lineno}: {e}")
        pass
    return_json = {'status': 'ok',
                'status_text': text, 'screenshot': file_name}

    return file_name



def closePM(row, self_call=False):
    global SEARCH_KEY
    if row['DATE-PM'] == 'nan' or row['DATE-PM'] == '':
        return {"status": 'ok', 'text': 'No PM Work', 'nosave': True}
    if row['PM-CLOSED'] != 'nan':
        return {"status": 'done', 'text': row['PM-CLOSED'], 'nosave': True}
    global emp_list
    if cookies['PHPSESSID'] is None:
        set_login()
        return closePM(row, self_call)
    response = False
    try:
        response = requests.get(
            "https://nsmart.nhealth-asia.com/MTDPDB01/pm/maintain_list.php?s_byear=" +
            row['YEAR'] + '&s_jobdate=' + formatDate(row['START-PLAN']) +
            '&s_to_date=' +
            formatDate(row['END-PLAN']) + '&' + SEARCH_KEY[1] + '=' + row[SEARCH_KEY[0]],
            headers=headers,
            cookies=cookies,
            verify=False,
        )
    except requests.exceptions.RequestException as e:
        print(f"An error occurred on line {sys.exc_info()[-1].tb_lineno}: {e}")
        time.sleep(5)
        return closePM(row, self_call)
    response.encoding = "tis-620"
    soup = BeautifulSoup(response.text, "lxml")
    # print(soup)
    table = soup.find("table", {"class", "Grid"})
    if table == None:
        print("No table found, re-login...")
        set_login()
        return closePM(row, self_call)
    tr = table.find('tr', {"class", "Row"})
    if tr == None:
        return {"status": 'fail', 'text': 'PM Work not found'}
    a_href = tr.find('a')['href'].split('?')[1]
    job_no = a_href.split('jobno=')[1].split('&')[0]
    if emp_list[row['TEAM'].lower()] is None or emp_list[row['TEAM'].lower()][row['ENGINEER'].lower()] is None:
        print('Team not found')
        get_emp_list()
        dept_tech = calibrator_list[row['TEAM'].lower()]
    else:
        dept_tech = calibrator_list[row['TEAM'].lower()]
    tool = ""
    if tools_list.get(confdata['SITE'].lower()) is None or tools_list.get(confdata['SITE'].lower()) == {}:
        load_tool_list(a_href)
        tool = tools_list[confdata['SITE'].lower()][row['TESTER'].lower()]
    else:
        tool = tools_list[confdata['SITE'].lower()][row['TESTER'].lower()]
    selects = [
        {'name': 'job_result',  'value': '1', 'text_contain': False},
        {'name': 'dept_tech',  'value': dept_tech, 'text_contain': False},
        {'name': 'toolid', 'value': tool, 'text_contain': False},
        {'name': 'app_issue_name',
            'value': row['INSPECTOR ID'], 'text_contain': False},
    ]
    date = formatDate(row['DATE-PM'])
    if date is not None and date != 'nan' and date != 'Invalid date' and date != '':
        inputs = [
            {'name': 'assign_date', 'value': date},
            {'name': 'act_dstart', 'value': date},
            {'name': 'act_dfin', 'value': date},
            {'name': 'approve_date', 'value': formatDate(row['ISSUE-PM'])}
        ]
    else:
        today = time.strftime("%d/%m/%Y")
        inputs = [
            {'name': 'assign_date', 'value': today},
            {'name': 'act_dstart', 'value': today},
            {'name': 'act_dfin', 'value': today},
            {'name': 'approve_date', 'value': formatDate(row['ISSUE-PM'])}
        ]
    #  create formdata foor post request
    form_data = {}
    for input in inputs:
        form_data[input['name']] = input['value']
    for select in selects:
        form_data[select['name']] = select['value']

    # find element in emp_vender that index [1] contain vender
    # if vender.startswith('[[CAL]]'):
    #     form_data['emp_id'] = 'ฺB005'.encode('tis-620')
    # else:
    #     emp_id = list(filter(lambda x: vender.lower()
    #                   in x[1].lower(), emp_list))
    #     if len(emp_id) == 0:
    #         get_emp_list(1)
    #         emp_id = list(filter(lambda x: vender.lower()
    #                       in x[1].lower(), emp_list))
    #     if len(emp_id) == 0:
    #         print('Vender not found')
    #         return
    #     form_data['emp_id'] = emp_id[0][0]
    form_data['emp_id'] = emp_list[row['TEAM'].lower()][row['ENGINEER'].lower()]
    if row['PM-STATUS'].lower() == 'pass':
        form_data['pass_status'] = '1'
    else:
        form_data['pass_status'] = '0'
    # return form_data
    response = requests.post('https://nsmart.nhealth-asia.com/MTDPDB01/pm/maintain07.php?' + a_href +
                            '&ccsForm=main_jobs%3AEdit', headers=headers, cookies=cookies, data=form_data, verify=False)
    response.encoding = "tis-620"
    soup = BeautifulSoup(response.text, "lxml")
    result_table = soup.find('table', {'class': 'Record'})
    if result_table == None:
        return {"status": 'fail', 'text': 'Fail to close PM'}
    result_tr = result_table.find('tr', {'class': 'Total'})
    result_td = result_tr.find('td')
    if result_td.text.strip() == 'PM status : Completed-send equipment back':
        return_json = 'SUCCESS'
        global screenshot
        if screenshot:
            file_name = get_screen_shot(
                soup, 'close_pm_css.css', result_td.text.strip())
            # move file to folder
            shutil.move(file_name+'.png', os.path.join(
                root_dir, "SCREENSHOT", "PM", row[SEARCH_KEY[0]]+"_"+job_no+".png"))
        # if self_call:
        #     return return_json
        return {"status": 'ok', 'text': result_td.text.strip()}


def closeCAL(row, self_call=False):
    global SEARCH_KEY
    if row['DATE-CAL'] == 'nan' or row['DATE-CAL'] == '':
        return {"status": 'ok', 'text': 'No CAL Work', 'nosave': True}
    if row['CAL-CLOSED'] != 'nan':
        return {"status": 'done', 'text': row['CAL-CLOSED'], 'nosave': True}
    if cookies['PHPSESSID'] is None:
        set_login()
        return closeCAL(row, self_call)
    response = False
    try:
        response = requests.get(
            "https://nsmart.nhealth-asia.com/MTDPDB01/caliber/caliber03.php?s_byear=" +
            row['YEAR'] + '&s_jobdate=' + formatDate(row['START-PLAN']) +
            '&s_to_date=' +
            formatDate(row['END-PLAN']) + '&' + SEARCH_KEY[1] + '=' + row[SEARCH_KEY[0]],
            headers=headers,
            cookies=cookies,
            verify=False,
        )
    except requests.exceptions.RequestException as e:
        print(f"An error occurred on line {sys.exc_info()[-1].tb_lineno}: {e}")
        time.sleep(5)
        return closeCAL(row, self_call)
    response.encoding = "tis-620"
    soup = BeautifulSoup(response.text, "lxml")
    # print(soup)
    table = soup.find("table", {"class", "Grid"})
    if table == None:
        print("No table found, re-login...")
        set_login()
        return closeCAL(row, self_call)
    tr = table.find('tr', {"class", "Row"})
    if tr == None:
        return {"status": 'fail', 'text': 'CAL Work not found'}
    a_href = tr.find('a')['href'].split('?')[1]

    job_no = a_href.split('jobno=')[1].split('&')[0]
    dept_caliber = calibrator_list[row['TEAM'].lower()]
    selects = [
        {'name': 'tech_idea_stat',  'value': '4', 'text_contain': False},
        {'name': 'dept_caliber',  'value': dept_caliber, 'text_contain': False},
    ]
    date = formatDate(row['DATE-CAL'])
    if date is not None and date != 'nan' and date != 'Invalid date' and date != '':
        inputs = [
            {'name': 'assign_date', 'value': date},
            {'name': 'act_dstart', 'value': date},
            {'name': 'act_dfin', 'value': date},
        ]
    else:
        today = time.strftime("%d/%m/%Y")
        inputs = [
            {'name': 'assign_date', 'value': today},
            {'name': 'act_dstart', 'value': today},
            {'name': 'act_dfin', 'value': today},
        ]

    form_data = {}
    emp_id = ''
    form_data['emp_id'] = emp_list[row['TEAM'].lower()][row['ENGINEER'].lower()]
    form_data['inspec_app_name'] = row['INSPECTOR NAME']
    if row['CAL-STATUS'].lower() == 'pass':
        form_data['CheckBox2'] = '1'
    else:
        form_data['CheckBox2'] = '0'
    for input in inputs:
        form_data[input['name']] = input['value']
    for select in selects:
        form_data[select['name']] = select['value']
    response = requests.post('https://nsmart.nhealth-asia.com/MTDPDB01/caliber/caliber03_1.php?' + a_href +
                            '&ccsForm=caliber_jobs_tech%3AEdit', headers=headers, cookies=cookies, data=form_data, verify=False)
    response.encoding = "tis-620"
    soup = BeautifulSoup(response.text, "lxml")
    result_td = list(filter(lambda x: 'Completed-send equipment back' in x.text.strip(), soup.find_all(
        'table', {'class': 'Record'})[1].find_all('tr', {'class': 'Controls'})[0].find_all('td')))

    if result_td == None or len(result_td) == 0:
        return {"status": 'fail', 'text': 'Fail to close CAL'}
    result_td = result_td[0]
    global screenshot
    if screenshot:
        file_name = get_screen_shot(
            soup, 'close_cal_css.css', result_td.text.strip())
        # move file to folder
        shutil.move(file_name+'.png', os.path.join(
            root_dir, "SCREENSHOT", "CAL", row[SEARCH_KEY[0]]+"_"+job_no+'.png'))
    return {"status": 'ok', 'text': "CAL status : "+result_td.text.strip()}


def attachFilePM(file_name_list, row):
# def attachFilePM(id, team, engineer, date, file_name_list, status, attach):
    global SEARCH_KEY
    id = row[SEARCH_KEY[0]]
    team = row['TEAM']
    engineer = row['ENGINEER']
    date = row['DATE-PM']
    status = row['PM-STATUS']
    attach = row['ATTACH-FILE-PM']
    if attach != 'yes':
        return {"status": 'ok', 'text': 'No need to attach PM file', 'nosave': True}
    if status.lower() == 'success':
        return {"status": 'done', 'text': 'Already attached PM file', 'nosave': True}
    start_date, end_date, now_year = getFirstAndLastDay(date)
    findname = [ele for ele in file_name_list if row['ID CODE']+'_'+now_year+'_pm' in ele]
    if len(findname) == 0:
        return {"status": 'fail', 'text': 'PM Work not found'}
    report_name = findname[0]
    # print type of dataurl
    if cookies['PHPSESSID'] is None:
        set_login()
        attachFilePM(file_name_list, row)
    response = False
    code = id
    try:
        response = requests.get(
            "https://nsmart.nhealth-asia.com/MTDPDB01/pm/maintain_list.php?s_byear=" +
            str(now_year) + '&s_jobdate=' + start_date +
            '&s_to_date=' + end_date + '&' + SEARCH_KEY[1] + '=' + code,
            headers=headers,
            cookies=cookies,
            verify=False,
        )
    except requests.exceptions.RequestException as e:
        print(f"An error occurred on line {sys.exc_info()[-1].tb_lineno}: {e}")
        time.sleep(5)
        return attachFilePM(file_name_list, row)
    response.encoding = "tis-620"
    soup = BeautifulSoup(response.text, "lxml")
    # print(soup)
    table = soup.find("table", {"class", "Grid"})
    if table == None:
        print("No table found, re-login...")
        set_login()
        return attachFilePM(file_name_list, row)
    tr = table.find('tr', {"class", "Row"})
    if tr == None:
        return {"status": 'fail', 'text': 'PM Work not found'}
    a_href = tr.find('a')['href'].split('?')[1]
    job_no = a_href.split('jobno=')[1].split('&')[0]
    response = requests.get(
        "https://nsmart.nhealth-asia.com/MTDPDB01/pm/maintain08.php?" +
        a_href + "&ccsForm=Maindocattache1",
        headers=headers,
        cookies=cookies,
        verify=False,
    )
    response.encoding = "tis-620"
    soup = BeautifulSoup(response.text, "lxml")
    form = soup.find('form', {'name': 'Post'})
    file_count = 0
    file_tr = soup.find_all('table', {'class': 'Grid'})[
        1].find_all('tr', {'class': 'Row'})
    for tr in file_tr:
        if tr.find_all('td')[2].text.strip() == 'Report PM ' + engineer.upper():
            file_count += 1

    if file_count > 0:
        return {"status": 'ok', 'text': 'Already attached PM file'}
    form_data = {}
    inputs = form.findAll('input')
    emp_name = ''
    for input in inputs:
        try:
            form_data[input['name']] = input['value']
        except:
            continue

    emp_name = team.upper()

    inputs = [
        {'name': 'docno', 'value': file_count+1},
        {'name': 'description_doc', 'value': 'Report PM ' + engineer.upper()},
        {'name': 'jobno', 'value': a_href.split('jobno=')[1].split('&')[0]}
    ]
    #  create formdata foor post request
    for input in inputs:
        form_data[input['name']] = input['value']

    # get file base64 from report folder
    file = open(os.path.join(root_dir, 'REPORTS',
                report_name + '.pdf'), 'rb').read()
    files = {'document_File': ('report.pdf', file, 'application/pdf')}

    # dataurl = json.loads(dataurl)
    # if dataurl.get('isImage') == True:
    #     dataurl = base64_image_to_base64_pdf(dataurl['base64'])
    # else:
    #     dataurl = dataurl['base64']
    # file = base64.b64decode(dataurl)
    # files = {'document_File': ('report.pdf', file, 'application/pdf')}

    # set form multipart/form-data
    # copy headers
    use_headers = headers.copy()
    # # use_headers['enctype'] = 'multipart/form-data'
    # use_headers['Content-Type'] = 'multipart/form-data'
    # maintain08.php?s_byear=2023&s_jobdate=&s_to_date=&s_pay=&s_job_status=&s_job_result=&s_branchid=&s_dept=&s_sub_dept=&s_code=&s_sap_code=1235&s_classno=&s_groupid=&s_catagory=&s_tpriority=&s_brand=&s_model=&s_serial_no=&s_inplan=&s_pmok=&maintain_list_vPageSize=&s_dept_tech=&s_sup_serv=&s_jobno=&s_docok=&code=36565&deptco=&jobno=1278631&ccsForm=Maindocattache1

    response2 = requests.post('https://nsmart.nhealth-asia.com/MTDPDB01/pm/maintain08.php?' + a_href +
                            '&ccsForm=Maindocattache1', headers=use_headers, cookies=cookies, data=form_data, files=files, verify=False)
    response2.encoding = "tis-620"
    soup = BeautifulSoup(response2.text, "lxml")
    result_table = soup.find_all('table', {'class': 'Header'})[1]
    if result_table == None:
        return {"status": 'fail', 'text': 'Fail to Attach PM file'}
    result_th = result_table.find('th')
    if result_th.text.strip() != 'Total : 0 records':
        global screenshot
        if screenshot:
            file_name = get_screen_shot(
                soup, 'close_pm_css.css', result_th.text.strip())
            # move file to folder
            shutil.move(file_name+'.png', os.path.join(
                root_dir, "SCREENSHOT", "PM", "Attach_PM_" + id+"_"+job_no+".png"))
        return {"status": 'ok', 'text': 'Attach PM file : '+result_th.text.strip()}
    else:
        return {"status": 'fail', 'text': 'Fail to Attach PM file'}


def attachFileCAL(file_name_list, row):
    global SEARCH_KEY
    id = row[SEARCH_KEY[0]]
    team = row['TEAM']
    engineer = row['ENGINEER']
    date = row['DATE-CAL']
    status = row['CAL-STATUS']
    attach = row['ATTACH-FILE-CAL']
    if attach != 'yes':
        return {"status": 'ok', 'text': 'No need to attach CAL file', 'nosave': True}
    if status.lower() == 'success':
        return {"status": 'done', 'text': 'Already attached CAL file', 'nosave': True}
    start_date, end_date, now_year = getFirstAndLastDay(date)
    findname = [ele for ele in file_name_list if row['ID CODE']+"_"+now_year+"_cal" in ele]
    if len(findname) == 0:
        return {"status": 'fail', 'text': 'CAL Work not found'}
    report_name = findname[0]
    if cookies['PHPSESSID'] is None:
        set_login()
        attachFileCAL(file_name_list, row)
    response = False
    code = id
    try:
        response = requests.get(
            "https://nsmart.nhealth-asia.com/MTDPDB01/caliber/caliber03.php?s_byear=" +
            str(now_year) + '&s_jobdate=' + start_date +
            '&s_to_date=' + end_date + '&' + SEARCH_KEY[1] + '=' + code,
            headers=headers,
            cookies=cookies,
            verify=False,
        )
    except requests.exceptions.RequestException as e:
        print(f"An error occurred on line {sys.exc_info()[-1].tb_lineno}: {e}")
        time.sleep(5)
        return attachFileCAL(file_name_list, row)
    response.encoding = "tis-620"
    soup = BeautifulSoup(response.text, "lxml")
    # print(soup)
    table = soup.find("table", {"class", "Grid"})
    if table == None:
        print("No table found, re-login...")
        set_login()
        return attachFileCAL(file_name_list, row)
    tr = table.find('tr', {"class", "Row"})
    if tr == None:
        return {"status": 'fail', 'text': 'CAL Work not found'}
    a_href = tr.find('a')['href'].split('?')[1]
    job_no = a_href.split('jobno=')[1].split('&')[0]
    response = requests.get(
        "https://nsmart.nhealth-asia.com/MTDPDB01/caliber/caliber03_5.php?" +
        a_href + "&ccsForm=caliber_jobs_tech%3AEdit",
        headers=headers,
        cookies=cookies,
        verify=False,
    )
    response.encoding = "tis-620"
    soup = BeautifulSoup(response.text, "lxml")
    form = soup.find('form', {'name': 'Post'})
    file_count = 0
    file_tr = soup.find_all('table', {'class': 'Grid'})[
        1].find_all('tr', {'class': 'Row'})
    for tr in file_tr:
        if tr.find_all('td')[2].text.strip() == 'Report CAL ' + engineer.upper():
            file_count += 1
    if file_count > 0:
        return {"status": 'ok', 'text': 'Already attached CAL file'}
    form_data = {}
    inputs = form.findAll('input')

    inputs = [
        {'name': 'docno', 'value': file_count+1},
        {'name': 'description_doc', 'value': 'Report CAL ' + engineer.upper()},
        {'name': 'jobno', 'value': a_href.split('jobno=')[1].split('&')[0]}
    ]
    #  create formdata foor post request
    for input in inputs:
        form_data[input['name']] = input['value']

    file = open(os.path.join(root_dir, 'REPORTS',
                report_name+'.pdf'), 'rb').read()
    files = {'document_File': ('report.pdf', file, 'application/pdf')}
    # copy headers
    use_headers = headers.copy()
    # # use_headers['enctype'] = 'multipart/form-data'
    # use_headers['Content-Type'] = 'multipart/form-data'
    # maintain08.php?s_byear=2023&s_jobdate=&s_to_date=&s_pay=&s_job_status=&s_job_result=&s_branchid=&s_dept=&s_sub_dept=&s_code=&s_sap_code=1235&s_classno=&s_groupid=&s_catagory=&s_tpriority=&s_brand=&s_model=&s_serial_no=&s_inplan=&s_pmok=&maintain_list_vPageSize=&s_dept_tech=&s_sup_serv=&s_jobno=&s_docok=&code=36565&deptco=&jobno=1278631&ccsForm=Maindocattache1

    response2 = requests.post("https://nsmart.nhealth-asia.com/MTDPDB01/caliber/caliber03_5.php?" + a_href +
                            "&ccsForm=Caliberdocattache1", headers=use_headers, cookies=cookies, data=form_data, files=files, verify=False)
    response2.encoding = "tis-620"
    soup = BeautifulSoup(response2.text, "lxml")
    result_table = soup.find_all('table', {'class': 'Header'})[1]
    if result_table == None:
        return {"status": 'fail', 'text': 'Fail to Attach CAL file'}
    result_th = result_table.find('th')
    if result_th.text.strip() != 'Total : 0 records':
        global screenshot
        if screenshot:
            file_name = get_screen_shot(
                soup, 'close_cal_css.css', result_th.text.strip())
            # move file to folder
            shutil.move(file_name+'.png', os.path.join(
                root_dir, "SCREENSHOT", "CAL", "Attach_CAL_" + id+"_"+job_no+".png"))
        return {"status": 'ok', 'text': 'Attach CAL file : '+result_th.text.strip()}
    else:
        return {"status": 'fail', 'text': 'Fail to Attach CAL file'}


screenshot = False




def read_file(option=None):
    global SEARCH_KEY
    print('[yellow]Select the SEARCH field[/yellow] [light blue][1]ID CODE[/light blue] [light blue][2]Item no.[/light blue] : ',end='')
    SEARCH_KEY = input()
    if SEARCH_KEY == '1':
        SEARCH_KEY = ['ID CODE', 's_sap_code']
    elif SEARCH_KEY == '2':
        SEARCH_KEY = ['ITEM NO', 's_code']
    df = read_excel_file()
    dir_path = ''
    if getattr(sys, 'frozen', False):
        dir_path = os.path.dirname(sys.executable)
    else:
        dir_path = os.path.dirname(os.path.realpath(__file__))
    path = os.path.join(dir_path, 'REPORTS')
    dir_list = os.listdir(path)
    file_name_list = ['xx']
    for file in dir_list:
        if file.endswith('.pdf'):
            file_name_list.append(file.replace('.pdf', ''))
    try:
        # read file
        print('[red]กำลังอ่านไฟล์...[/red]')

        with pd.ExcelWriter(os.path.join(excel_folder, excel_file_name).replace('//', '/'),
                            mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
            shutil.copy(os.path.join(excel_folder, excel_file_name).replace('//', '/'),
                        os.path.join(excel_folder, 'backup' + excel_file_name).replace('//', '/'))
            # count row['PM-STATUS] to lower case = 'pass'
            pass_pm = 0
            pass_cal = 0
            attach_pm = 0
            attach_cal = 0

            for index, row in df.iterrows():
                if row['PM-CLOSED'].lower() == 'nan':
                    pass_pm += 1
                if row['CAL-CLOSED'].lower() == 'nan':
                    pass_cal += 1
                if row['PM-ATTACH-STATUS'].lower() == 'nan':
                    attach_pm += 1
                if row['CAL-ATTACH-STATUS'].lower() == 'nan':
                    attach_cal += 1
            table = Table(title='Total Medical Devices Found: {} devices'.format(
                len(df)), title_justify='center', title_style='bold magenta')
            table.add_column('#', justify='left', style='cyan', no_wrap=True)
            table.add_column('Number of Devices', justify='right',
                            style='green', no_wrap=True)
            if (option == 'close_pm_cal' or option == None):
                table.add_row('PM Jobs Not Closed', str(pass_pm))
                table.add_row('CAL Jobs Not Closed', str(pass_cal))

            if (option == 'attach_pm_cal' or option == None):
                table.add_row('PM Jobs Not Attached', str(attach_pm))
                table.add_row('CAL Jobs Not Attached', str(attach_cal))

            Console().print(table)
            print(f'\n[red]Do you want to start the process? (Y/N): [/red]', end='')
            start = input()
            if start.lower() == 'y':
                print(
                    f'\n[red]Do you want to save screenshots? (Y/N): [/red]', end='')
                global screenshot
                screenshot = input().lower() == 'y'
                init_text = pyfiglet.figlet_format("Start Process...")
                print(init_text)
                start_time = time.time()
                total_time = 0
                average_time = 5
                estimated_end_time = 0
                try:
                    with Progress() as progress:
                        task = progress.add_task("[red]Processing...[/red]", total=len(df))
                        for index, row in df.iterrows():
                            row[SEARCH_KEY[0]] = str(row[SEARCH_KEY[0]]).replace('.0', '')
                            row['YEAR'] = str(row['YEAR']).replace('.0', '')
                            print('[green]{}[/green] / [blue]{}[/blue] [yellow]Close PM JOB[/yellow] [light blue]{}[/light blue]'.format(
                                str(index+1), str(len(df)), row[SEARCH_KEY[0]])
                            )
                            if row[SEARCH_KEY[0]] != 'nan':
                                issave = False
                                process_result_pm = ''
                                process_result_cal = ''
                                process_result_pm_attach = ''
                                process_result_cal_attach = ''
                                # future = executor.submit(foo, 'world!')
                                if option == 'close_pm_cal':
                                    with ThreadPoolExecutor() as executor:
                                        process_result_pm = executor.submit(
                                            closePM, row)
                                        process_result_cal = executor.submit(
                                            closeCAL, row)
                                    master_df['PM-CLOSED'] = master_df['PM-CLOSED'].astype(
                                        str)
                                    master_df['CAL-CLOSED'] = master_df['CAL-CLOSED'].astype(
                                        str)
                                    process_result_pm = process_result_pm.result()
                                    process_result_cal = process_result_cal.result()
                                    # if process_result_pm == 'SUCCESS' or process_result_pm == '':
                                    
                                    if process_result_pm.get('status') == 'ok' or process_result_pm.get('status') == 'done':
                                        issave = False if issave == False and process_result_pm.get('nosave') == True else True
                                        df.at[index, 'PM-CLOSED'] = "SUCCESS" if "Completed" in process_result_pm.get('text') else process_result_pm.get('text')
                                        if process_result_pm.get('status') == 'ok':
                                            print('[green]{}[/green]'.format(process_result_pm.get('text')))
                                    else:
                                        print('[red]{}[/red]'.format(process_result_pm.get('text')))
                                        df.at[index, 'PM-CLOSED'] = process_result_pm.get('text')
                                    if process_result_cal.get('status') == 'ok' or process_result_cal.get('status') == 'done':
                                        issave = False if issave == False and process_result_cal.get('nosave') == True else True
                                        df.at[index, 'CAL-CLOSED'] = "SUCCESS" if "Completed" in process_result_cal.get('text') else process_result_cal.get('text')
                                        if process_result_cal.get('status') == 'ok':
                                            print('[green]{}[/green]'.format(process_result_cal.get('text')))
                                    elif process_result_cal.get('status') != 'done':
                                        print('[red]{}[/red]'.format(process_result_cal.get('text')))
                                        df.at[index, 'CAL-CLOSED'] = process_result_cal.get('text')
                                    master_df['PM-ATTACH-STATUS'] = master_df['PM-ATTACH-STATUS'].astype(
                                        str)
                                    master_df['CAL-ATTACH-STATUS'] = master_df['CAL-ATTACH-STATUS'].astype(
                                        str)
                                    df.at[index, 'PM-ATTACH-STATUS'] = row['PM-ATTACH-STATUS'] if row['PM-ATTACH-STATUS'] != 'nan' else ''
                                    df.at[index, 'CAL-ATTACH-STATUS'] = row['CAL-ATTACH-STATUS'] if row['CAL-ATTACH-STATUS'] != 'nan' else ''

                                elif option == 'attach_pm_cal':
                                    with ThreadPoolExecutor() as executor:
                                        process_result_pm = executor.submit(
                                            attachFilePM, file_name_list, row)
                                        process_result_cal = executor.submit(
                                            attachFileCAL, file_name_list, row)
                                    master_df['PM-ATTACH-STATUS'] = master_df['PM-ATTACH-STATUS'].astype(
                                        str)
                                    master_df['CAL-ATTACH-STATUS'] = master_df['CAL-ATTACH-STATUS'].astype(
                                        str)
                                    process_result_pm = process_result_pm.result()
                                    process_result_cal = process_result_cal.result()
                                    if process_result_pm.get('status') == 'ok' or process_result_pm.get('status') == 'done':
                                        issave = False if issave == False and process_result_pm.get('nosave') == True else True
                                        df.at[index, 'PM-ATTACH-STATUS'] = 'SUCCESS' if process_result_pm.get('text').startswith('Attach') else process_result_pm.get('text')
                                        if process_result_pm.get('status') == 'ok':
                                            print('[green]{}[/green]'.format(process_result_pm.get('text')))
                                    else:
                                        print('[red]{}[/red]'.format(process_result_pm.get('text')))
                                        df.at[index, 'PM-ATTACH-STATUS'] = process_result_pm.get('text')
                                    if process_result_cal.get('status') == 'ok' or process_result_cal.get('status') == 'done':
                                        issave = False if issave == False and process_result_pm.get('nosave') == True else True
                                        if process_result_cal.get('status') == 'ok':
                                            df.at[index, 'CAL-ATTACH-STATUS'] = 'SUCCESS' if process_result_cal.get('text').startswith('Attach') else process_result_cal.get('text')
                                        print('[green]{}[/green]'.format(process_result_cal.get('text')))
                                    else:
                                        print('[red]{}[/red]'.format(process_result_cal.get('text')))
                                        df.at[index, 'CAL-ATTACH-STATUS'] = process_result_cal.get('text')
                                    master_df['PM-CLOSED'] = master_df['PM-CLOSED'].astype(
                                        str)
                                    master_df['CAL-CLOSED'] = master_df['CAL-CLOSED'].astype(
                                        str)
                                    df.at[index, 'PM-CLOSED'] = row['PM-CLOSED'] if row['PM-CLOSED'] != 'nan' else ''
                                    df.at[index, 'CAL-CLOSED'] = row['CAL-CLOSED'] if row['CAL-CLOSED'] != 'nan' else ''
                                elif option == None:
                                    process_result_pm = ''
                                    process_result_cal = ''
                                    with ThreadPoolExecutor() as executor:
                                        process_result_pm = executor.submit(
                                            closePM, row)
                                        process_result_cal = executor.submit(
                                            closeCAL, row)
                                        process_result_pm_attach = executor.submit(
                                            attachFilePM, file_name_list, row)
                                        process_result_cal_attach = executor.submit(
                                            attachFileCAL, file_name_list, row)
                                    master_df['PM-CLOSED'] = master_df['PM-CLOSED'].astype(
                                        str)
                                    master_df['CAL-CLOSED'] = master_df['CAL-CLOSED'].astype(
                                        str)
                                    master_df['PM-ATTACH-STATUS'] = master_df['PM-ATTACH-STATUS'].astype(
                                        str)
                                    master_df['CAL-ATTACH-STATUS'] = master_df['CAL-ATTACH-STATUS'].astype(
                                        str)
                                    process_result_pm = process_result_pm.result()
                                    process_result_cal = process_result_cal.result()
                                    process_result_pm_attach = process_result_pm_attach.result()
                                    process_result_cal_attach = process_result_cal_attach.result()
                                    if process_result_pm.get('status') == 'ok' or process_result_pm.get('status') == 'done':
                                        issave = False if issave == False and process_result_pm.get('nosave') == True else True
                                        df.at[index, 'PM-CLOSED'] = "SUCCESS" if "Completed" in process_result_pm.get('text') else process_result_pm.get('text')
                                        if process_result_pm.get('status') == 'ok':
                                            print('[green]{}[/green]'.format(process_result_pm.get('text')))
                                    else:
                                        print('[red]{}[/red]'.format(process_result_pm.get('text')))
                                        df.at[index, 'PM-CLOSED'] = process_result_pm.get('text')
                                    if process_result_cal.get('status') == 'ok' or process_result_cal.get('status') == 'done':
                                        issave = False if issave == False and process_result_pm.get('nosave') == True else True
                                        df.at[index, 'CAL-CLOSED'] = "SUCCESS" if "Completed" in process_result_cal.get('text') else process_result_cal.get('text')
                                        if process_result_cal.get('status') == 'ok':
                                            print('[green]{}[/green]'.format(process_result_cal.get('text')))
                                    else:
                                        print('[red]{}[/red]'.format(process_result_cal.get('text')))
                                        df.at[index, 'CAL-CLOSED'] = process_result_cal.get('text')
                                    if process_result_pm_attach.get('status') == 'ok' or process_result_pm_attach.get('status') == 'done':
                                        issave = False if issave == False and process_result_pm.get('nosave') == True else True
                                        df.at[index, 'PM-ATTACH-STATUS'] = "SUCCESS" if process_result_pm_attach.get('text').startswith('Attach') else process_result_pm_attach.get('text')
                                        if process_result_pm_attach.get('status') == 'ok':
                                            print('[green]{}[/green]'.format(process_result_pm_attach.get('text')))
                                    else:
                                        print('[red]{}[/red]'.format(process_result_pm_attach.get('text')))
                                        df.at[index, 'PM-ATTACH-STATUS'] = process_result_pm_attach.get('text')
                                        issave = False if issave == False and process_result_pm.get('nosave') == True else True
                                    if process_result_cal_attach.get('status') == 'ok' or process_result_cal_attach.get('status') == 'done':
                                        issave = False if issave == False and process_result_pm.get('nosave') == True else True
                                        df.at[index, 'CAL-ATTACH-STATUS'] = "SUCCESS" if process_result_cal_attach.get('text').startswith('Attach') else process_result_cal_attach.get('text')
                                        if process_result_cal_attach.get('status') == 'ok':
                                            print('[green]{}[/green]'.format(process_result_cal_attach.get('text')))
                                    else:
                                        print('[red]{}[/red]'.format(process_result_cal_attach.get('text')))
                                        df.at[index, 'CAL-ATTACH-STATUS'] = process_result_cal_attach.get('text')

                                if (issave and index != 0 and index % 20 == 0) or index == len(df)-1:
                                    # convert 'nan' to empty string with match entire cell
                                    print(
                                        '[green]===========Auto Saved=============[/green]')
                                    df.style.format(lambda v: "number-format: @").to_excel(
                                        writer, sheet_name='Sheet1', index=False)
                                    issave = False
                            progress.update(task, advance=1)
                    
                except Exception as e:
                    print(
                        f"An error occurred on line {sys.exc_info()[-1].tb_lineno}: {e}")
                finally:
                    print(
                        '[green]=============Save Result to Excel==============[/green]')
                    # convert 'nan' to empty string with match entire cell
                    df = df.replace('^nan$', '', regex=True)
                    # format excel to string
                    df.style.format(lambda v: "number-format: @").to_excel(
                        writer, sheet_name='Sheet1', index=False)
                    # save df to excel
                    writer.close()

                end_time = time.time()
                print('\n[green]Closed jobs and attached files successfully[/green]')
                print('[green]Completed in[/green]: [yellow]{}[/yellow] seconds'.format(
                    convertTime(end_time - start_time)))
                print('Press any button to return to the main menu: ', end='')
                input()
                # clear console
                os.system('cls' if os.name == 'nt' else 'clear')
                init_text = pyfiglet.figlet_format(
                    "BME Assistant", font="slant")
                print(init_text)
                showmenu()
            else:
                print("[red]Cancel[/red]")
                # clear console
                os.system('cls' if os.name == 'nt' else 'clear')
                init_text = pyfiglet.figlet_format(
                    "BME Assistant", font="slant")
                print(init_text)
                showmenu()
                # pyautogui.alert('Close Jobs '+str(index)+' records')
            writer.close()
    except Exception as e:
        print(f"An error occurred on line {sys.exc_info()[-1].tb_lineno}: {e}")
        print('[red]Error[/red]')
        print('Press any button to return to the main menu: ', end='')
        input()
        # clear console
        os.system('cls' if os.name == 'nt' else 'clear')
        init_text = pyfiglet.figlet_format("BME Assistant", font="slant")
        print(init_text)
        showmenu()

def convertDate(date):
    if date is None or len(date) == 0 or date == '-' or date[0] == '':
        return ""
    date = date[0].split(':')[1].strip()
    months_str = ['JANUARY', 'FEBRUARY', 'MARCH', 'APRIL', 'MAY', 'JUNE',
                    'JULY', 'AUGUST', 'SEPTEMBER', 'OCTOBER', 'NOVEMBER', 'DECEMBER']
    date = date.split(' ')
    if len(date) == 1:
        return ""
    if len(date) == 3:
        month = months_str.index(date[1].upper()) + 1
        return date[0] + '/' + str(month) + '/' + date[2]
    return ' '.join(date)

def getYear(date1, date2):
    if date1 == '' and date2 == '':
        return ''
    date = ''
    if date1 == '':
        date = date2
    elif date2 == '':
        date = date1
    else:
        date = date1
    date = date.split('/')
    return date[2]

def getStartMonth(date1, date2):
    # convert date string to date object
    if date1 == '' and date2 == '':
        return ''
    date = ''
    if date1 == '':
        date = date2
    elif date2 == '':
        date = date1
    else:
        date = date1
    # get first day of month
    date = date.split('/')
    date[0] = '1'
    date_object = datetime.strptime('/'.join(date), '%d/%m/%Y')
    return date_object.strftime('%d/%m/%Y')

def getEndMonth(date1, date2):
    # convert date string to date object
    if date1 == '' and date2 == '':
        return ''
    date = ''
    if date1 == '':
        date = date2
    elif date2 == '':
        date = date1
    else:
        date = date1
    # get last day of month
    date = date.split('/')
    date[0] = str(calendar.monthrange(int(date[2]), int(date[1]))[1])
    date_object = datetime.strptime('/'.join(date), '%d/%m/%Y')
    return date_object.strftime('%d/%m/%Y')

def getTeamName(engineer):
    global emp_list
    engineer = engineer.lower()
    if emp_list is None or len(emp_list) == 0:
        load_empList()
    # find team name in emp_list
    for team in emp_list:
        if engineer in emp_list[team]:
            return emp_list[team]['option_name']
    return ''

def change_file_name():
    dir_path = ''
    if getattr(sys, 'frozen', False):
        dir_path = os.path.dirname(sys.executable)
    else:
        dir_path = os.path.dirname(os.path.realpath(__file__))
    report_path = os.path.join(dir_path, 'REPORTS')
    # check if path has subfolder
    subfolder = os.listdir(report_path)

    def move_file_to_root(folder):
        subfolder = os.listdir(folder)
        for file in subfolder:
            if os.path.isdir(os.path.join(folder, file)):
                move_file_to_root(os.path.join(folder, file))
            else:
                # move file to root folder
                shutil.move(os.path.join(folder, file),
                            os.path.join(report_path, file))
        os.rmdir(folder)

    if len(subfolder) > 0:
        # move file to root folder
        for file in subfolder:
            if os.path.isdir(os.path.join(report_path, file)):
                move_file_to_root(os.path.join(report_path, file))
    dir_list = os.listdir(report_path)
    name_arr = {}
    with Progress() as progress:
        task = progress.add_task(
            "[cyan]Renaming files[/cyan]", total=len(dir_list))
        for file_name in dir_list:
            source = os.path.join(report_path, file_name)
            # if file is not pdf
            if not file_name.endswith('.pdf'):
                print(
                    '[red]Not a PDF file[/red] : [yellow]{}[/yellow]'.format(file_name))
                progress.update(task, advance=1)
                continue

            # if file_name.find('_') == -1:
            file = fitz.open(source)
            page = file[0]
            # find text with regex /ID CODE.*\n.*$/gm
            page = page.get_text()
            # replace text in page with regex /^-\n/gm
            page = re.sub(r'^-\n', '', page, flags=re.MULTILINE)
            text = re.findall(r'ID CODE.*\n.*$', page, re.MULTILINE)
            if len(text) == 0:
                print(
                    '[red]No equipment code found in PDF[/red] : [yellow]{}[/yellow]'.format(file_name))
                progress.update(task, advance=1)
                continue
            code = text[0].split('\n')[1].replace(':', '').strip()
            if code.find('(') > -1:
                code = code.split('(')[0].strip()
            cal = re.findall(r'Certificate', page, re.MULTILINE)
            year = None
            if len(cal) > 0:
                caldate = re.findall(
                    r'CALIBRATED DATE.*\n.*$', page, re.MULTILINE)
                issuedate = re.findall(
                    r'ISSUE DATE.*\n.*$', page, re.MULTILINE)
                
                year = getYear(convertDate(caldate), convertDate(issuedate))
                if name_arr.get(code+'#'+year) is None:
                    name_arr[code+'#'+year] = {}
                name_arr[code+'#'+year]['cal'] = caldate
                name_arr[code+'#'+year]['issue-cal'] = issuedate
                name = code + "_"+year + '_cal.pdf'
            else:
                pmdate = re.findall(r'PM. DATE.*\n.*$', page, re.MULTILINE)
                issuedate = re.findall(
                    r'ISSUE DATE.*\n.*$', page, re.MULTILINE)
                
                year = getYear(convertDate(pmdate), convertDate(issuedate))
                if name_arr.get(code+'#'+year) is None:
                    name_arr[code+'#'+year] = {}
                name_arr[code+'#'+year]['pm'] = pmdate
                name_arr[code+'#'+year]['issue-pm'] = issuedate
                name = code + "_"+year+'_pm.pdf'
            safety = re.findall(
                r'electricalsafetyanalyzer', page.lower().replace('\n', '').replace(' ',''), re.MULTILINE)
            if safety is not None and len(safety) > 0:
                name_arr[code+'#'+year]['safety'] = 'Electrical Safety Analyzer'
            else:
                name_arr[code+'#'+year]['safety'] = '-'
            engineer = re.findall(r'Approved by.*\n.*$', page, re.MULTILINE)
            if engineer is not None and len(engineer) > 0:
                engineer = engineer[0].replace('\n', '').split(':')[
                    1].strip().replace('  ', ' ')
                name_arr[code+'#'+year]['engineer'] = engineer

            else:
                name_arr[code + '#'+year]['engineer'] = '-'
            # file.save(os.path.join(path, name))
            department = re.findall(r'LOCATION.*\n.*$', page, re.MULTILINE)
            if department is not None and len(department) > 0:
                department = department[0].replace('\n', '').split(':')[
                    1].strip().replace('  ', ' ')
                name_arr[code + '#'+year]['department'] = department
            file.close()
            try:
                os.rename(source, os.path.join(report_path, name))
            except Exception as e:
                # if already exist
                os.remove(os.path.join(report_path, name))
                os.rename(source, os.path.join(report_path, name))
            # writer = PdfWriter(clone_from=os.path.join(report_path, name))
            # for page in writer.pages:
            #     page.compress_content_streams(level=9)
            # with open(os.path.join(report_path, name), 'wb') as f:
            #     writer.write(f)

            print('[grey42]{}[/grey42] [yellow]>>>>[/yellow] [green]{}[/green]'.format(
                file_name, name))
            progress.update(task, advance=1)
            # else:
            #     filename = file_name.split('_')
            #     if len(filename) > 1 and len(filename[1]) > 10:
            #         filename = "_".join(filename[1:])
            #         filename = re.sub(r'\s\(\d{1,}\)', '', filename)
            #         name = filename.split('(')[0]
            #         if (filename.split('(')[-1] != '1).pdf' and filename.split('(')[-1] != '2).pdf'):
            #             name_arr.append(name)
            #             if filename.split('(')[1].find('PM') > -1:
            #                 name = name + '_pm.pdf'
            #             else:
            #                 name = name + '_cal.pdf'
            #             print('เปลี่ยนชื่อไฟล์ [yellow]{}[/yellow] เป็น [yellow]{}[/yellow]'.format(
            #                 file_name, name))
            #             dest = os.path.join(path, name)
            #             os.rename(source, dest)

            # append name_arr to clipboard

            # convert name_arr to table
        # create array with length is 20 and fill with ''
        unique_arr = []
        for i, x in enumerate(name_arr):
            id = x
            value = name_arr[x]
            if value.get('cal') is None:
                value['cal'] = ['']
            if value.get('pm') is None:
                value['pm'] = ['']
            tmp_arr = [''] * 26
            tmp_arr[0] = str(i+1)
            tmp_arr[1] = id.split('#')[0]
            tmp_arr[2] = getTeamName(value['engineer'])
            tmp_arr[7] = convertDate(value['pm'])

            if tmp_arr[7] == '':
                tmp_arr[8] = ''
                tmp_arr[13] = ''
                tmp_arr[15] = ''
                tmp_arr[18] = ''
            else:
                tmp_arr[8] = convertDate(value['issue-pm'])
                if tmp_arr[8] == '' or tmp_arr[8] == '-':
                    tmp_arr[8] = convertDate(value['pm'])
                tmp_arr[13] = 'PM doable'
                tmp_arr[15] = 'pass'
                tmp_arr[18] = 'yes'
            tmp_arr[9] = convertDate(value['cal'])
            if tmp_arr[9] == '':
                tmp_arr[10] = ''
                tmp_arr[14] = ''
                tmp_arr[16] = ''
                tmp_arr[19] = ''
            else:
                tmp_arr[10] = convertDate(value['issue-cal'])
                if tmp_arr[10] == '' or tmp_arr[10] == '-':
                    tmp_arr[10] = convertDate(value['cal'])
                tmp_arr[14] = 'Perform CAL'
                tmp_arr[16] = 'pass'
                tmp_arr[19] = 'yes'
            tmp_arr[4] = id.split('#')[1]
            tmp_arr[5] = getStartMonth(tmp_arr[7], tmp_arr[9])
            tmp_arr[6] = getEndMonth(tmp_arr[7], tmp_arr[9])
            tmp_arr[11] = confdata["SUP_ID"]
            tmp_arr[12] = confdata["SUP_NAME"]
            tmp_arr[17] = value['safety']
            tmp_arr[3] = value['engineer']
            tmp_arr[24] = value['department']
            unique_arr.append(tmp_arr)

            # if value.get('cal') is not None and value.get('pm') is not None:
            #     cal = value.get('cal')[0].split('\n')[1].replace(':', '').strip()
            #     pm = value.get('pm')[0].split('\n')[1].replace(':', '').strip()
            #     name_arr[id] = {'cal': cal, 'pm': pm}
            # else:
            #     name_arr[id] = {'cal': 'nan', 'pm': 'nan'}
        # copy to cilpboard to paste in excel

        line_strings = []
        for line in unique_arr:
            line_strings.append('\t'.join(line).replace('\n', ''))
        arr_string = '\r\n'.join(line_strings)
        pyperclip.copy(arr_string)
        print('\n[green]Successfully copied equipment codes[/green]')
        print('\n[green]File name change completed[/green]')
        print('\n[purple]Press any key to return to the main menu: [/purple]', end='')
        input()
    # clear console
    os.system('cls' if os.name == 'nt' else 'clear')
    print(init_text)
    showmenu()

def re_init_app():
    today = datetime.now()
    confdata['Last run'] = today.strftime('%d/%m/%Y')
    with open(os.path.join(root_dir, 'CONFIG', 'config.json'), 'w') as f:
        json.dump(confdata, f)
    # delete file calibarator_list.json
    for file in ['calibrator_list.json', 'tool_list.json', 'emp_list.json']:
        if os.path.exists(os.path.join(root_dir, 'CONFIG', file)):
            os.remove(os.path.join(
                root_dir, 'CONFIG', file))
    load_empList()
    load_calibrator_list()
    load_tool_list('none')
    os.system('cls' if os.name == 'nt' else 'clear')
    print(init_text)
   
        


def showmenu():
    last_run_date = confdata.get('Last run')
    today = datetime.now()
    if (last_run_date != today.strftime('%d/%m/%Y')):
        re_init_app()
    # menu
    Console().print('[light blue]Welcome to the job closing program[/light blue]')
    print('[light blue]Please select a menu[/light blue]')
    print('[yellow][1] Open script for downloading ECERT files[/yellow]')
    print('[yellow][2] Change file name[/yellow]')
    print('[yellow][3][/yellow] [red]Close[/red] [yellow]PM and CAL[/yellow]')
    print('[yellow][4][/yellow] [red]Attach[/red] [yellow]PM and CAL files[/yellow]')
    print('[yellow][5][/yellow] [red]Close[/red] [yellow]and[/yellow] [red]Attach[/red] [yellow]PM and CAL files[/yellow]')
    print('[white]----------------------------------------[/white]')
    print('[red]Additional Programs[/red]')
    print('[yellow][6] Get all equipment data[/yellow]')
    print('[yellow][7] Reinitialize Application[/yellow]')
    # set input color to blue
    print(f'\n[purple]Press a number to select the menu: [/purple]', end='')
    menu = input()
    if menu == '1':
        dir_path = ''
        if getattr(sys, 'frozen', False):
            dir_path = os.path.dirname(sys.executable)
        else:
            dir_path = os.path.dirname(os.path.realpath(__file__))
        # open file "download cert.txt" at SOURCE folder in notepad
        # os.system('notepad.exe "' + os.path.join(dir_path,
        #           'SOURCE', 'download cert.txt') + '"')
        # print('[green]เปิดสคริปต์สำหรับดาวน์โหลดไฟล์ ECERT[/green]')
        script_text = open(os.path.join(dir_path, 'SOURCE',
                        'download cert.txt'), encoding='utf-8').read()
        # copy to clipboard
        pyperclip.copy(script_text)
        print('\n[green]Script for downloading ECERT files copied successfully[/green]')
        print('\nPress any key to return to the main menu: ', end='')
        input()
        os.system('cls' if os.name == 'nt' else 'clear')
        print(init_text)
        showmenu()
    elif menu == '2':
        change_file_name()
        showmenu()
    elif menu == '6':
        get_equipment_file()
        showmenu()
    
    elif menu == '7':
        re_init_app()
        showmenu()
    else:
        load_empList()
        load_calibrator_list()
        load_tool_list('none')

        if menu == '3':
            read_file('close_pm_cal')
        elif menu == '4':
            read_file('attach_pm_cal')
        elif menu == '5':
            read_file()

        else:
            print('Menu not found')
            showmenu()


try:
    showmenu()
except Exception as e:
    print(f"An error occurred on line {sys.exc_info()[-1].tb_lineno}: {e}")
# change_file_name()
# get_emp_list()
# get_equipment_file()
