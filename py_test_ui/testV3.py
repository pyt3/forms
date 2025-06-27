import base64
from io import BytesIO
import re
import numpy as np
import openpyxl
import pandas as pd
import os
from PIL import Image
from html2image import Html2Image
import requests
from bs4 import BeautifulSoup
import winreg
import threading
from datetime import datetime
import json
from concurrent.futures import ThreadPoolExecutor
import pyfiglet
import urllib3
from urllib.parse import urlparse
from queue import Queue
import time
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
from rich import print, pretty
from rich.progress import track, Progress, BarColumn
from rich.table import Table
from rich.console import Console
from rich import box
from rich.panel import Panel
from rich.align import Align
import sys
import logging
import shutil
import calendar
import pyperclip
import fitz
from pypdf import PdfReader, PdfWriter
import random

# # set cmd to support utf-8
# os.system('chcp 874')
# os.system('$LANG="th_TH.UTF-8"')
# os.system('cls')

# from closePM import closePM
# from closeCAL import closeCAL

from openpyxl import load_workbook


class LineLimitRotatingFileHandler(logging.Handler):
    def __init__(self, filename, max_lines=50000):
        super().__init__()
        self.filename = filename
        self.max_lines = max_lines
        self._ensure_log_file_exists()

    def _ensure_log_file_exists(self):
        """Ensure the log file exists, creating it if necessary."""
        if not os.path.exists(self.filename):
            with open(self.filename, 'w'):
                pass

    def _get_line_count(self):
        """Return the number of lines in the log file."""
        with open(self.filename, 'r') as f:
            return len(f.readlines())

    def _rotate_log(self):
        """Rotate the log file to limit the number of lines."""
        with open(self.filename, 'r') as f:
            lines = f.readlines()

        # Keep the last max_lines lines
        with open(self.filename, 'w') as f:
            f.writelines(lines[-self.max_lines:])

    def emit(self, record):
        """Write the log record to the log file and check line count."""
        try:
            # Check the current line count
            if self._get_line_count() >= self.max_lines:
                self._rotate_log()

            # Write the log message to the file
            with open(self.filename, 'a') as f:
                f.write(self.format(record) + '\n')
        except Exception:
            self.handleError(record)


logger = logging.getLogger('my_logger')
logger.setLevel(logging.ERROR)
handler = LineLimitRotatingFileHandler('log.txt', max_lines=50000)
formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
handler.setFormatter(formatter)
logger.addHandler(handler)
# logging.basicConfig(filename='log.txt', level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
# logging.basicConfig(level=logging.ERROR)

# set depth of recursion
sys.setrecursionlimit(10000)

# pretty.install()
# pretty.pretty_print = True

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

init_text = pyfiglet.figlet_format("CES Assistant", font="slant")
print(init_text)
root_dir = os.path.dirname(os.path.abspath(__file__))

# check if using pyinstaller
if getattr(sys, 'frozen', False):
    root_dir = os.path.dirname(sys.executable)
global config
global confdata
global login_site
global data
global cookies
config = open(os.path.join(root_dir, "CONFIG", "config.json"), "r")
confdata = json.load(config)

login_site = confdata["SITE"]

lock = threading.Lock()
data = {
    "user": confdata["USERNAME"],
    "pass": confdata["PASSWORD"],
    "Submit": "Submit"
}

# The above Python code snippet defines a dictionary named `cookies` containing three key-value pairs
# representing cookies used in a web request. The keys represent the cookie names, and the values are
# the corresponding cookie values.
cookies = {
    "_ga": "GA1.1.409909798.1624509466",
    "_ga_L9CPT990SV": "GS1.1.1629863162.2.0.1629863162.0",
    "PHPSESSID": confdata["SESSION_ID"],
}


headers = {
    "Connection": "keep-alive",
    "Upgrade-Insecure-Requests": "1",
    "User-Agent": "Mozilla/5.0 (Windows NT 6.1; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/99.0.4844.51 Safari/537.36",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9",
    "Referer": "https://nsmart.nhealth-asia.com/mtdpdb01/default.php",
    "Accept-Language": "th-TH,th;q=0.9,en;q=0.8",
}
emp_list = {}
tools_list = {}
calibrator_list = {}
temp_emp_list = {}
SEARCH_KEY = 's_code'

excel_file_name = 'recordCal_PM.xlsx'
excel_folder = os.path.join(root_dir, 'EXCEL FILE')


def read_excel_file():
    global master_df
    global SEARCH_KEY
    master_df = (pd.read_excel(
        os.path.join(excel_folder, excel_file_name).replace('//', '/'), sheet_name="Sheet1", engine='openpyxl')).dropna(subset=[SEARCH_KEY[0]])
    # check if dataframe gasattr map
    if not hasattr(master_df, 'map'):
        master_df.map = master_df.applymap
    return master_df.map(str)


def internet_connection():
    try:
        requests.get('https://www.google.com/', verify=False)
        return True
    except requests.exceptions.RequestException as e:
        return False


def set_login(max_retries=3, initial_delay=1):
    """Establish a login session with the remote server and update session IDs.
    
    Args:
        max_retries: Maximum number of retry attempts (default: 3)
        initial_delay: Initial delay between retries in seconds (default: 1)
    
    Returns:
        bool: True if login was successful, False otherwise
    """
    global confdata, cookies
    
    retry_count = 0
    delay = initial_delay
    
    while retry_count < max_retries:
        try:
            with requests.Session() as s:
                r = s.post(
                    "https://nsmart.nhealth-asia.com/MTDPDB01/index.php",
                    data=data,
                    headers=headers,
                    verify=False,
                    timeout=10  # Add timeout to prevent hanging
                )
                
                # Check if login was successful (status code or response content check)
                if r.status_code == 200 and 'PHPSESSID' in s.cookies:
                    # Update session information
                    confdata["SESSION_ID"] = s.cookies['PHPSESSID']
                    cookies['PHPSESSID'] = s.cookies['PHPSESSID']
                    
                    # Save updated configuration
                    config_path = os.path.join(root_dir, "CONFIG/config.json")
                    with open(config_path, "w") as json_file:
                        json.dump(confdata, json_file)
                    
                    return True
                else:
                    print(f"[yellow]Login failed with status code: {r.status_code}[/yellow]")
            
            # Increment retry counter and delay
            retry_count += 1
            time.sleep(delay)
            delay *= 2  # Exponential backoff
            
        except requests.exceptions.RequestException as e:
            print(f"[red]Network error on attempt {retry_count+1}/{max_retries}: {e}[/red]")
            retry_count += 1
            if retry_count < max_retries:
                print(f"[yellow]Retrying in {delay} seconds...[/yellow]")
                time.sleep(delay)
                delay *= 2  # Exponential backoff
        except Exception as e:
            print(f"[red]Unexpected error on line {sys.exc_info()[-1].tb_lineno}: {e}[/red]")
            retry_count += 1
            if retry_count < max_retries:
                print(f"[yellow]Retrying in {delay} seconds...[/yellow]")
                time.sleep(delay)
                delay *= 2  # Exponential backoff
    
    # If we've exhausted all retries
    print("[red]Failed to login after multiple attempts. Check your internet connection or credentials.[/red]")
    return False


def formatDate(date):
    if date != 'nan':
        # date is dd/mm/yyyy
        date = date.split('/')
        date[0] = '0' + date[0] if len(date[0]) == 1 else date[0]
        date[1] = '0' + date[1] if len(date[1]) == 1 else date[1]
        date = '/'.join(date)
    return date


def convertTime(time):
    if time != 'nan':
        # convert time is second to text format
        hours = time / 3600
        minutes = (time % 3600) / 60
        seconds = time % 60
        time = f'{int(hours)} hours {int(minutes)} minutes {int(seconds)} seconds'
        return time

    return time


def save_empList(emp_list=None):
    # write over file to clear old data
    if emp_list is None:
        return
    with open(os.path.join(root_dir, 'CONFIG/emp_list.json'), 'w') as json_file:
        json.dump(emp_list, json_file)


def save_calibrator_list(calibrator_list=None):
    """Save calibrator list to JSON file."""
    if calibrator_list is None:
        return
    config_path = os.path.join(root_dir, 'CONFIG', 'calibrator_list.json')
    try:
        with open(config_path, 'w') as json_file:
            json.dump(calibrator_list, json_file)
    except Exception as e:
        logging.error(f"Failed to save calibrator list: {e}")


def load_empList():
    """Load employee list from JSON file or fetch if not available."""
    global emp_list
    try:
        emp_list_path = os.path.join(root_dir, 'CONFIG', 'emp_list.json')
        if os.path.exists(emp_list_path):
            with open(emp_list_path, 'r') as json_file:
                emp_list = json.load(json_file)
            if not emp_list:
                print("[yellow]Employee list is empty, fetching from server...[/yellow]")
                get_emp_list()
        else:
            print("[yellow]Employee list not found, fetching from server...[/yellow]")
            get_emp_list()
    except Exception as e:
        logging.error(f"Error loading employee list: {e}")
        print(f"[red]Error loading employee list: {e}[/red]")
        get_emp_list()


def load_tool_list(href):
    """Load tool list from JSON file or fetch if not available."""
    global tools_list
    
    # Fetch new tools if href is provided
    if href is not None and href != 'none':
        try:
            print('[yellow]Loading tools list from server...[/yellow]')
            response = requests.get(
                f"https://nsmart.nhealth-asia.com/MTDPDB01/pm/maintain07.php?{href}&ccsForm=main_jobs%3AEdit",
                headers=headers,
                cookies=cookies,
                verify=False,
                timeout=10
            )
            response.raise_for_status()
            response.encoding = "tis-620"
            
            soup = BeautifulSoup(response.text, "lxml")
            select = soup.find('select', {'name': 'toolid'})
            
            if select and len(select) > 0:
                site_key = confdata['SITE'].lower()
                if tools_list is None:
                    tools_list = {}
                if site_key not in tools_list:
                    tools_list[site_key] = {}
                
                for option in select.find_all('option'):
                    tools_list[site_key][option.text.strip().lower()] = option['value']
                
                save_tool_list(tools_list)
                print('[green]Tools list updated successfully[/green]')
            else:
                print('[red]No tools found in the response[/red]')
        except requests.exceptions.RequestException as e:
            print(f"[red]Network error while loading tools: {e}[/red]")
            time.sleep(3)
            return load_tool_list(href)
        except Exception as e:
            print(f"[red]Error parsing tools list: {e}[/red]")
    
    # Load from cache file
    try:
        tools_list_path = os.path.join(root_dir, 'CONFIG', 'tools_list.json')
        if os.path.exists(tools_list_path):
            with open(tools_list_path, 'r') as json_file:
                tools_list = json.load(json_file)
        
        if not tools_list:
            tools_list = {}
            save_tool_list(tools_list)
    except Exception as e:
        logging.error(f"Error loading tools list: {e}")
        tools_list = {}
        save_tool_list(tools_list)


def save_tool_list(tools_list=None):
    """Save tools list to JSON file."""
    if tools_list is None:
        return
    tools_path = os.path.join(root_dir, 'CONFIG', 'tools_list.json')
    try:
        with open(tools_path, 'w') as json_file:
            json.dump(tools_list, json_file)
    except Exception as e:
        logging.error(f"Failed to save tools list: {e}")


def load_calibrator_list():
    """Load calibrator list from JSON file or fetch if not available."""
    global calibrator_list
    try:
        calibrator_path = os.path.join(root_dir, 'CONFIG', 'calibrator_list.json')
        if os.path.exists(calibrator_path):
            with open(calibrator_path, 'r') as json_file:
                calibrator_list = json.load(json_file)
            
            if not calibrator_list:
                print('[yellow]Calibrator list is empty, fetching from server...[/yellow]')
                get_emp_list()
        else:
            print('[yellow]Calibrator list not found, fetching from server...[/yellow]')
            get_emp_list()
    except Exception as e:
        logging.error(f"Error loading calibrator list: {e}")
        print(f"[red]Error loading calibrator list: {e}[/red]")
        get_emp_list()


equipments_arr = []


def get_equipment_file(url='https://nsmart.nhealth-asia.com/MTDPDB01/asset_mast_list_new.php?asset_masterPageSize=100', page='1', max_retries=3):
    """Fetch equipment list from server and save to Excel file.
    
    Args:
        url: Base URL for fetching equipment data
        page: Current page number
        max_retries: Maximum number of retry attempts
    
    Returns:
        List of equipment data
    """
    global equipments_arr
    
    # Initialize array for first page
    if page == '1':
        equipments_arr = []
        print('[cyan]Starting equipment data collection...[/cyan]')
    
    # Construct page URL
    page_url = f"{url}&asset_masterPage={page}"
    
    # Fetch data with retry logic
    for retry in range(max_retries):
        try:
            response = requests.get(
                page_url,
                headers=headers,
                cookies=cookies,
                verify=False,
                timeout=15
            )
            response.raise_for_status()
            response.encoding = "tis-620"
            
            # Find table with equipment data
            tables = re.findall(r'<table\s+[^>]*class=["\']Grid["\'][^>]*>', response.text, re.DOTALL)
            if not tables:
                print("[yellow]No data table found, re-authenticating...[/yellow]")
                if set_login():
                    return get_equipment_file(url, page)
                else:
                    print("[red]Authentication failed, cannot continue[/red]")
                    return equipments_arr
            
            # Extract table content
            table_content = response.text.split('<table class="Grid" cellspacing="0" cellpadding="0">')[1]
            table_content = table_content.split('</table>')[0]
            table_html = '<table class="Grid" cellspacing="0" cellpadding="0">' + table_content + '</table>'
            
            # Parse table
            soup = BeautifulSoup(table_html, "lxml")
            footer = soup.find('tr', {'class': 'Footer'})
            
            if not footer:
                print("[red]Invalid page structure, cannot extract data[/red]")
                return equipments_arr
                
            # Get max page number
            max_page = footer.text.split('of')[1].strip().split(' ')[0]
            print(f'[yellow]Fetching page[/yellow] [blue]{page}[/blue] [yellow]out of[/yellow] [blue]{max_page}[/blue]')
            
            # Extract rows
            rows = soup.find_all('tr')
            
            # Extract headers from first page
            header = [ele.text.strip() for ele in rows[1].find_all('th')]
            
            # Skip header row on subsequent pages
            data_rows = rows[1:-1] if page == '1' else rows[2:-1]
            
            # Process rows
            for row in data_rows:
                cols = row.find_all('td')
                
                # Extract image link if available
                img_link = ''
                if cols[1].find('a') is not None:
                    img_link = 'https://nsmart.nhealth-asia.com/MTDPDB01/' + cols[1].find('a').get('href')
                
                # Extract text from columns
                cols_data = [ele.text.strip() if ele.text.strip() != 'Click' else '' for ele in cols]
                cols_data[1] = img_link
                
                # Add to results
                equipments_arr.append(cols_data)
            
            # Check if we've reached the last page
            if int(page) >= int(max_page):
                # Save to Excel
                excel_path = os.path.join(root_dir, 'EXCEL FILE', 'equipment_list.xlsx')
                
                with Console().status("[bold green]Saving equipment data to Excel...[/bold green]"):
                    df = pd.DataFrame(equipments_arr, columns=header)
                    df.to_excel(excel_path, index=False)
                    
                    # Auto-adjust column widths
                    auto_adjust_column_width_from_df(excel_path, 'Sheet1')
                
                print(f'[green]Successfully saved[/green] [blue]{len(equipments_arr)}[/blue] [green]equipment records[/green]')
                print(f'[yellow]Equipment list saved at:[/yellow] [blue]{excel_path}[/blue]')
                
                # Open the file
                os.system(f'start excel.exe "{excel_path}"')
                
                return equipments_arr
            else:
                # Fetch next page
                return get_equipment_file(url, str(int(page) + 1))
                
        except requests.exceptions.RequestException as e:
            print(f"[red]Network error on attempt {retry+1}/{max_retries}: {e}[/red]")
            if retry < max_retries - 1:
                wait_time = 2 * (retry + 1)  # Exponential backoff
                print(f"[yellow]Retrying in {wait_time} seconds...[/yellow]")
                time.sleep(wait_time)
            else:
                print("[red]Failed to fetch equipment data after multiple attempts[/red]")
                return equipments_arr
                
        except Exception as e:
            print(f"[red]Error processing equipment data on line {sys.exc_info()[-1].tb_lineno}: {e}[/red]")
            logging.error(f"Error in get_equipment_file: {e}")
            if retry < max_retries - 1:
                wait_time = 2 * (retry + 1)
                print(f"[yellow]Retrying in {wait_time} seconds...[/yellow]")
                time.sleep(wait_time)
            else:
                print("[red]Failed to process equipment data after multiple attempts[/red]")
                return equipments_arr

def auto_adjust_column_width_from_df(file_path, sheet_name):
    # Load the workbook and select the specified sheet
    workbook = load_workbook(file_path)
    worksheet = workbook[sheet_name]

    for column in worksheet.columns:
        max_length = 0
        for cell in column:
            try:
                # Check the length of the cell value
                cell_length = len(str(cell.value))
                max_length = max(max_length, cell_length)
            except:
                pass
        # Adjust the column width; add a little extra space
        adjusted_width = (max_length + 2)
        worksheet.column_dimensions[openpyxl.utils.get_column_letter(
            column[0].column)].width = adjusted_width

    # Save the workbook
    workbook.save(file_path)


temp_team_list = {}


def get_team_list(url, page='1', max_retries=3):
    """Fetch team member list from server with pagination support.
    
    Args:
        url: Base URL for team data
        page: Current page number
        max_retries: Maximum retry attempts
        
    Returns:
        Dictionary of team members with names (lowercase) as keys and IDs as values
    """
    global temp_team_list
    
    # Initialize empty dictionary for first page
    if page == '1':
        temp_team_list = {}
        print('[yellow]Fetching team members...[/yellow]')
    
    # Construct page URL
    page_url = f"{url}&employeePage={page}"
    
    # Try request with retry logic
    for retry in range(max_retries):
        try:
            response = requests.get(
                page_url,
                headers=headers,
                cookies=cookies,
                verify=False,
                timeout=10  # Add timeout to prevent hanging
            )
            response.raise_for_status()
            response.encoding = "tis-620"
            
            soup = BeautifulSoup(response.text, "lxml")
            form = soup.find('form', {'name': 'employee'})
            
            # Check if form exists
            if form is None:
                print("[yellow]Authentication required, logging in...[/yellow]")
                if set_login():
                    return get_team_list(url, page)
                else:
                    print("[red]Login failed, could not fetch team list[/red]")
                    return temp_team_list
            
            # Get rows and footer
            rows = form.find_all('tr', {'class': 'Row'})
            footer_text = form.find('tr', {'class': 'Footer'}).text.strip()
            
            # Extract max page number
            max_page = footer_text.split('of')[1].strip().split(' ')[0]
            
            # Check if we've exceeded max pages
            if int(page) > int(max_page):
                return temp_team_list
            
            # Get expected total members
            max_emp = int(rows[0].find('td').text.strip().split('\xa0')[1])
            
            # Check if we already have all members
            if len(temp_team_list) >= max_emp:
                return temp_team_list
            
            # Process current page
            if max_emp > 0 and len(rows) > 1:
                # Skip header row
                for row in rows[1:]:
                    inputs = row.find_all('input')
                    if len(inputs) >= 2 and inputs[0]['value'] and inputs[1]['value']:
                        member_name = inputs[1]['value'].lower().replace('  ', ' ')
                        member_id = inputs[0]['value']
                        temp_team_list[member_name] = member_id
                
                # Fetch next page if not at the end
                if int(page) < int(max_page):
                    return get_team_list(url, str(int(page) + 1))
            
            return temp_team_list
            
        except requests.exceptions.RequestException as e:
            print(f"[red]Network error on attempt {retry+1}/{max_retries}: {e}[/red]")
            if retry < max_retries - 1:
                wait_time = 2 * (retry + 1)  # Exponential backoff
                print(f"[yellow]Retrying in {wait_time} seconds...[/yellow]")
                time.sleep(wait_time)
            else:
                print("[red]Failed to fetch team list after multiple attempts[/red]")
                return temp_team_list
                
        except Exception as e:
            print(f"[red]Error processing team data on line {sys.exc_info()[-1].tb_lineno}: {e}[/red]")
            logging.error(f"Error in get_team_list: {e}")
            if retry < max_retries - 1:
                wait_time = 2 * (retry + 1)
                print(f"[yellow]Retrying in {wait_time} seconds...[/yellow]")
                time.sleep(wait_time)
            else:
                print("[red]Failed to process team data after multiple attempts[/red]")
                return temp_team_list


def get_emp_list(max_retries=3):
    """Fetch all teams and their members from the server.
    
    Args:
        max_retries: Maximum number of retry attempts
        
    Returns:
        Updated global emp_list and calibrator_list dictionaries
    """
    global emp_list, calibrator_list
    
    # Initialize dictionaries if needed
    if not emp_list:
        emp_list = {}
    if not calibrator_list:
        calibrator_list = {}
    
    # Show status message
    print('[cyan]Fetching department information...[/cyan]')
    
    # Try request with retry logic
    for retry in range(max_retries):
        try:
            response = requests.get(
                "https://nsmart.nhealth-asia.com/MTDPDB01/reftable/employee_branch.php?dept_control=1",
                headers=headers,
                cookies=cookies,
                verify=False,
                timeout=10
            )
            response.raise_for_status()
            response.encoding = "tis-620"
            
            soup = BeautifulSoup(response.text, "lxml")
            form = soup.find('form', {'name': 'm_dept_tech'})
            
            if form is None:
                print("[yellow]Authentication required, logging in...[/yellow]")
                if set_login():
                    return get_emp_list()
                else:
                    print("[red]Login failed, could not fetch department list[/red]")
                    return
            
            # Process each department/team
            rows = form.find_all('tr', {'class': 'Row'})
            if len(rows) <= 1:
                print("[yellow]No departments found, check your permissions[/yellow]")
                return
            
            # Skip header row
            with Console().status("[bold green]Loading team data...[/bold green]"):
                for row in rows[1:]:
                    cells = row.find_all('td')
                    if len(cells) < 2:
                        continue
                        
                    team_id = cells[0].text.strip()
                    team_name = cells[1].text.strip()
                    team_name_lower = team_name.lower()
                    
                    # Store calibrator ID
                    calibrator_list[team_name_lower] = team_id
                    
                    # Fetch team members
                    print(f'[yellow]Fetching members of team[/yellow] [blue]{team_name}[/blue]')
                    url = f"https://nsmart.nhealth-asia.com/MTDPDB01/reftable/employee_branch.php?dept_control=1&dept_tech={team_id}"
                    team_members = get_team_list(url)
                    
                    # Store team data
                    emp_list[team_name_lower] = team_members
                    emp_list[team_name_lower]["option_name"] = team_name
            
            # Save data to files
            save_calibrator_list(calibrator_list)
            save_empList(emp_list)
            
            print("[green]Successfully fetched all team data[/green]")
            return
            
        except requests.exceptions.RequestException as e:
            print(f"[red]Network error on attempt {retry+1}/{max_retries}: {e}[/red]")
            if retry < max_retries - 1:
                wait_time = 2 * (retry + 1)  # Exponential backoff
                print(f"[yellow]Retrying in {wait_time} seconds...[/yellow]")
                time.sleep(wait_time)
            else:
                print("[red]Failed to fetch departments after multiple attempts[/red]")
                
        except Exception as e:
            print(f"[red]Error processing department data on line {sys.exc_info()[-1].tb_lineno}: {e}[/red]")
            logging.error(f"Error in get_emp_list: {e}")
            if retry < max_retries - 1:
                wait_time = 2 * (retry + 1)
                print(f"[yellow]Retrying in {wait_time} seconds...[/yellow]")
                time.sleep(wait_time)
            else:
                print("[red]Failed to process department data after multiple attempts[/red]")

def getFirstAndLastDay(date):
    if date is None or date == '':
        date = datetime.now()
    # check if date is string
    if isinstance(date, str):
        date = datetime.strptime(date, '%d/%m/%Y')
    # get first and last day of month
    first_day = date.replace(day=1)
    last_day = date.replace(day=calendar.monthrange(date.year, date.month)[1])
    first_day = first_day.strftime("%d/%m/%Y")
    last_day = last_day.strftime("%d/%m/%Y")
    date_year = first_day.split('/')[2]
    return first_day, last_day, date_year


def base64_image_to_base64_pdf(base64_image_string):
    try:
        # Decode the base64 image string to bytes
        image_data = base64.b64decode(base64_image_string)

        # Create a BytesIO object to open the image with Pillow
        image_buffer = BytesIO(image_data)

        # Open the image using Pillow
        image = Image.open(image_buffer)

        # Create a BytesIO object to save the PDF
        pdf_buffer = BytesIO()
        image.save(pdf_buffer, "PDF", resolution=100.0, save_all=True)

        # Get the content of the PDF as bytes
        pdf_bytes = pdf_buffer.getvalue()

        # Encode the PDF as a base64 string
        pdf_base64 = base64.b64encode(pdf_bytes).decode("utf-8")

        return pdf_base64

    except Exception as e:
        # print error with line number
        print(f"An error occurred on line {sys.exc_info()[-1].tb_lineno}: {e}")
        return None


def get_screen_shot(soup, css_file, text, type, code=""):
    # open css file from css folder
    css_file = open(os.path.join(root_dir, 'SOURCE/'+css_file), 'r')
    css = css_file.read()

    # get file path
    __location__ = os.path.join(root_dir, 'SCREENSHOT', type)
    # if folder not exist, create folder
    if os.path.exists(__location__) == False:
        # create folder
        os.makedirs(__location__)

    hti = Html2Image(output_path=__location__, size=(978, 675) if type == 'CAL' else (978, 475), disable_logging=True,
                     temp_path=__location__, custom_flags=['--default-background-color=ffffff', '--hide-scrollbars'])
    # filename is time in format 20240506
    file_name = datetime.now().strftime("%Y%m%d")
    html_str = str(soup).replace('src="../images/',
                                 'src="https://nsmart.nhealth-asia.com/MTDPDB01/images/')
    html_str = html_str.replace(
        'src="../Styles/', 'src="https://nsmart.nhealth-asia.com/MTDPDB01/Styles/')
    html_str = html_str.replace(
        'src="../img.php', 'src="https://nsmart.nhealth-asia.com/MTDPDB01/img.php')
    html_str = html_str.replace(
        'src="../files', 'src="https://nsmart.nhealth-asia.com/MTDPDB01/files')
    try:
        hti.screenshot(html_str=html_str, css_str=css,
                       save_as=str(code)+"-"+file_name+'.png')
        # convert to webp
        # im = Image.open(os.path.join(__location__, str(code)+"-"+file_name+'.png'))
        # im.save(os.path.join(__location__, str(code)+"-"+file_name+'.webp'), 'webp')

        # # remove png file
        # os.remove(os.path.join(__location__, str(code)+"-"+file_name+'.png'))
    except Exception as e:
        print(f"An error occurred on line {sys.exc_info()[-1].tb_lineno}: {e}")
        pass
    return_json = {'status': 'ok',
                   'status_text': text, 'screenshot': file_name}

    return file_name


def closePM(row, self_call=False):
    global SEARCH_KEY
    # Early returns for no-work or already-closed cases
    if row['DATE-PM'] == 'nan' or row['DATE-PM'] == '':
        return {"status": 'ok', 'text': 'No PM Work', 'nosave': True}
    if row['PM-CLOSED'].lower() == 'success':
        return {"status": 'done', 'text': row['PM-CLOSED'], 'nosave': True}
    
    # Check session and login if needed
    if cookies['PHPSESSID'] is None:
        set_login()
        return closePM(row, self_call)
    
    # Prepare URL with all parameters at once
    url = (f"https://nsmart.nhealth-asia.com/MTDPDB01/pm/maintain_list.php?s_byear={row['YEAR']}"
           f"&s_jobdate={formatDate(row['START-PLAN'])}&s_to_date={formatDate(row['END-PLAN'])}"
           f"&{SEARCH_KEY[1]}={row[SEARCH_KEY[0]]}")
    
    try:
        # Use timeout to prevent hanging requests
        response = requests.get(
            url, headers=headers, cookies=cookies, verify=False, timeout=10
        )
        response.raise_for_status()
        response.encoding = "tis-620"
        
        soup = BeautifulSoup(response.text, "lxml")
        table = soup.find("table", {"class", "Grid"})
        
        # Handle authentication or empty results
        if not table:
            set_login()
            return closePM(row, self_call)
            
        tr = table.find('tr', {"class", "Row"})
        if not tr:
            return {"status": 'fail', 'text': 'PM Work not found'}
        
        # Extract job details
        a_href = tr.find('a')['href'].split('?')[1]
        job_no = a_href.split('jobno=')[1].split('&')[0]
        
        # Get team and tool info efficiently
        team_lower = row['TEAM'].lower()
        if team_lower not in emp_list or row['ENGINEER'].lower() not in emp_list[team_lower]:
            get_emp_list()
            dept_tech = calibrator_list[team_lower]
        else:
            dept_tech = calibrator_list[team_lower]
        
        # Get tool information
        site_key = confdata['SITE'].lower()
        if site_key not in tools_list or not tools_list[site_key]:
            load_tool_list(a_href)
        tool = tools_list[site_key][row['TESTER'].lower()]
        
        # Prepare form data all at once
        form_data = {
            'job_result': '1',
            'dept_tech': dept_tech,
            'toolid': tool,
            'app_issue_name': row['INSPECTOR ID'],
            'emp_id': emp_list[team_lower][row['ENGINEER'].lower()],
            'pass_status': '1' if row['PM-STATUS'].lower() == 'pass' else '0'
        }
        
        # Set dates
        date = formatDate(row['DATE-PM'])
        if date and date != 'nan' and date != 'Invalid date' and date != '':
            form_data.update({
                'assign_date': date,
                'act_dstart': date,
                'act_dfin': date,
                'approve_date': formatDate(row['ISSUE-PM'])
            })
        else:
            today = time.strftime("%d/%m/%Y")
            form_data.update({
                'assign_date': today,
                'act_dstart': today,
                'act_dfin': today,
                'approve_date': formatDate(row['ISSUE-PM'])
            })
        
        # Submit form
        post_url = f'https://nsmart.nhealth-asia.com/MTDPDB01/pm/maintain07.php?{a_href}&ccsForm=main_jobs%3AEdit'
        response = requests.post(post_url, headers=headers, cookies=cookies, data=form_data, verify=False, timeout=10)
        response.raise_for_status()
        response.encoding = "tis-620"
        
        soup = BeautifulSoup(response.text, "lxml")
        result_table = soup.find('table', {'class': 'Record'})
        
        if not result_table:
            return {"status": 'fail', 'text': 'Fail to close PM'}
        
        result_tr = result_table.find('tr', {'class': 'Total'})
        result_td = result_tr.find('td')
        
        if result_td.text.strip() == 'PM status : Completed-send equipment back':
            # Take screenshot if enabled
            if screenshot:
                get_screen_shot(soup, 'close_pm_css.css', result_td.text.strip(), "PM", row[SEARCH_KEY[0]])
            
            return {"status": 'ok', 'text': result_td.text.strip()}
        return {"status": 'fail', 'text': 'Unknown PM status: ' + result_td.text.strip()}
            
    except requests.exceptions.RequestException as e:
        logging.error(f"Network error in closePM: {e}")
        time.sleep(3)  # Reduced sleep time
        return closePM(row, self_call)
    except Exception as e:
        logging.error(f"Error in closePM on line {sys.exc_info()[-1].tb_lineno}: {e}")
        return {"status": 'fail', 'text': f'Error: {str(e)}'}


def closeCAL(row, self_call=False):
    global SEARCH_KEY
    # Early returns for no-work or already-closed cases
    if row['DATE-CAL'] == 'nan' or row['DATE-CAL'] == '':
        return {"status": 'ok', 'text': 'No CAL Work', 'nosave': True}
    if row['CAL-CLOSED'].lower() == 'success':
        return {"status": 'done', 'text': row['CAL-CLOSED'], 'nosave': True}
    
    # Check session and login if needed
    if cookies['PHPSESSID'] is None:
        set_login()
        return closeCAL(row, self_call)
    
    # Prepare URL with all parameters at once
    url = (f"https://nsmart.nhealth-asia.com/MTDPDB01/caliber/caliber03.php?s_byear={row['YEAR']}"
           f"&s_jobdate={formatDate(row['START-PLAN'])}&s_to_date={formatDate(row['END-PLAN'])}"
           f"&{SEARCH_KEY[1]}={row[SEARCH_KEY[0]]}")
    
    try:
        # Use timeout to prevent hanging requests
        response = requests.get(
            url, headers=headers, cookies=cookies, verify=False, timeout=10
        )
        response.raise_for_status()
        response.encoding = "tis-620"
        
        soup = BeautifulSoup(response.text, "lxml")
        table = soup.find("table", {"class", "Grid"})
        
        # Handle authentication or empty results
        if not table:
            set_login()
            return closeCAL(row, self_call)
            
        tr = table.find('tr', {"class", "Row"})
        if not tr:
            return {"status": 'fail', 'text': 'CAL Work not found'}
        
        # Extract job details
        a_href = tr.find('a')['href'].split('?')[1]
        job_no = a_href.split('jobno=')[1].split('&')[0]
        
        # Get team info efficiently
        team_lower = row['TEAM'].lower()
        dept_caliber = calibrator_list[team_lower]
        
        # Prepare form data all at once
        form_data = {
            'tech_idea_stat': '4',
            'dept_caliber': dept_caliber,
            'emp_id': emp_list[team_lower][row['ENGINEER'].lower()],
            'inspec_app_name': row['INSPECTOR NAME'],
            'CheckBox2': '1' if row['CAL-STATUS'].lower() == 'pass' else '0'
        }
        
        # Set dates
        date = formatDate(row['DATE-CAL'])
        if date and date != 'nan' and date != 'Invalid date' and date != '':
            form_data.update({
                'assign_date': date,
                'act_dstart': date,
                'act_dfin': date
            })
        else:
            today = time.strftime("%d/%m/%Y")
            form_data.update({
                'assign_date': today,
                'act_dstart': today,
                'act_dfin': today
            })
        
        # Submit form
        post_url = f'https://nsmart.nhealth-asia.com/MTDPDB01/caliber/caliber03_1.php?{a_href}&ccsForm=caliber_jobs_tech%3AEdit'
        response = requests.post(post_url, headers=headers, cookies=cookies, data=form_data, verify=False, timeout=10)
        response.raise_for_status()
        response.encoding = "tis-620"
        
        soup = BeautifulSoup(response.text, "lxml")
        # More efficient way to find result text
        result_td = None
        try:
            records_table = soup.find_all('table', {'class': 'Record'})[1]
            controls_row = records_table.find('tr', {'class': 'Controls'})
            for td in controls_row.find_all('td'):
                if 'Completed-send equipment back' in td.text.strip():
                    result_td = td
                    break
        except (IndexError, AttributeError):
            pass
        
        if not result_td:
            return {"status": 'fail', 'text': 'Fail to close CAL'}
        
        # Take screenshot if enabled
        if screenshot:
            get_screen_shot(soup, 'close_cal_css.css', result_td.text.strip(), "CAL", row[SEARCH_KEY[0]])
        
        return {"status": 'ok', 'text': "CAL status : " + result_td.text.strip()}
            
    except requests.exceptions.RequestException as e:
        logging.error(f"Network error in closeCAL: {e}")
        time.sleep(3)  # Reduced sleep time
        return closeCAL(row, self_call)
    except Exception as e:
        logging.error(f"Error in closeCAL on line {sys.exc_info()[-1].tb_lineno}: {e}")
        return {"status": 'fail', 'text': f'Error: {str(e)}'}


def attachFilePM(file_name_list, row):
    global SEARCH_KEY
    # Early returns for no-work, already-attached, or no-need cases
    if row['DATE-PM'] == 'nan' or row['DATE-PM'] == '':
        return {"status": 'ok', 'text': 'No PM Work', 'nosave': True}
    if row['ATTACH-FILE-PM'] != 'yes':
        return {"status": 'ok', 'text': 'No need to attach PM file', 'nosave': True}
    if row['PM-ATTACH-STATUS'].lower() == 'success':
        return {"status": 'done', 'text': row['PM-ATTACH-STATUS'], 'nosave': True}
    
    # Check session and login if needed
    if cookies['PHPSESSID'] is None:
        set_login()
        return attachFilePM(file_name_list, row)
    
    # Find report file
    code = row[SEARCH_KEY[0]]
    now_year = row['YEAR']
    findname = [ele for ele in file_name_list if row['ID CODE'] + '_' + now_year + '_pm' in ele]
    if not findname:
        return {"status": 'fail', 'text': 'PM File For Attach not found'}
    report_name = findname[0]
    
    # Get job details
    start_date = formatDate(row['START-PLAN'])
    end_date = formatDate(row['END-PLAN'])
    
    # Build URL for request
    url = (f"https://nsmart.nhealth-asia.com/MTDPDB01/pm/maintain_list.php?s_byear={now_year}"
           f"&s_jobdate={start_date}&s_to_date={end_date}&{SEARCH_KEY[1]}={code}")
    
    try:
        # Use timeout to prevent hanging requests
        response = requests.get(
            url, headers=headers, cookies=cookies, verify=False, timeout=10
        )
        response.raise_for_status()
        response.encoding = "tis-620"
        
        soup = BeautifulSoup(response.text, "lxml")
        table = soup.find("table", {"class", "Grid"})
        
        # Handle authentication or empty results
        if not table:
            set_login()
            return attachFilePM(file_name_list, row)
            
        tr = table.find('tr', {"class", "Row"})
        if not tr:
            return {"status": 'fail', 'text': 'PM Work not found'}
        
        # Get attachment page
        a_href = tr.find('a')['href'].split('?')[1]
        attach_url = f"https://nsmart.nhealth-asia.com/MTDPDB01/pm/maintain08.php?{a_href}&ccsForm=Maindocattache1"
        response = requests.get(
            attach_url, headers=headers, cookies=cookies, verify=False, timeout=10
        )
        response.raise_for_status()
        response.encoding = "tis-620"
        
        soup = BeautifulSoup(response.text, "lxml")
        form = soup.find('form', {'name': 'Post'})
        
        if not form:
            set_login()
            return attachFilePM(file_name_list, row)
        
        # Find and delete existing files
        file_tr = soup.find_all('table', {'class': 'Grid'})[1].find_all('tr', {'class': 'Row'})
        file_count = []
        engineer = row['ENGINEER'].upper()
        
        for tr in file_tr:
            if tr.find_all('td')[2].text.strip() == 'Report PM ' + engineer:
                file_count.append(tr.find_all('td')[0].find('a')['href'])
        
        # Batch delete old files
        if file_count:
            for url, i in zip(file_count, range(len(file_count))):
                logging.info(f'Deleting old PM file: {i+1}/{len(file_count)} {code}')
                try:
                    # Get delete form
                    del_url = f"https://nsmart.nhealth-asia.com/MTDPDB01/pm/{url}"
                    del_response = requests.get(
                        del_url, headers=headers, cookies=cookies, verify=False, timeout=10
                    )
                    del_response.raise_for_status()
                    del_response.encoding = "tis-620"
                    
                    del_soup = BeautifulSoup(del_response.text, "lxml")
                    del_form = del_soup.find('form', {'name': 'Post'})
                    
                    if not del_form:
                        continue
                    
                    # Prepare delete form data
                    form_data = {}
                    for inp in del_form.findAll('input'):
                        try:
                            form_data[inp['name']] = inp['value']
                        except KeyError:
                            continue
                    
                    # Remove update button and add delete button coordinates
                    if "Button_Update" in form_data:
                        form_data.pop("Button_Update")
                    form_data['Button_Delete.x'] = str(random.randint(10, 50))
                    form_data['Button_Delete.y'] = str(random.randint(5, 20))
                    
                    # Submit delete request
                    del_post_url = f"https://nsmart.nhealth-asia.com/MTDPDB01/pm/{url}&ccsForm=Maindocattache1%3AEdit"
                    del_response = requests.post(
                        del_post_url, headers=headers, cookies=cookies, data=form_data, verify=False, timeout=10
                    )
                    del_response.raise_for_status()
                    
                except requests.exceptions.RequestException as e:
                    logging.error(f"Error deleting PM file: {e}")
                    time.sleep(2)  # Shorter sleep time
            
            print(f'[red]Delete old PM file: {len(file_count)} file(s) success[/red] [yellow]{code}[/yellow]')
        
        # Prepare form data for file upload
        form_data = {}
        for inp in form.findAll('input'):
            try:
                form_data[inp['name']] = inp['value']
            except KeyError:
                continue
        
        # Add required fields
        form_data.update({
            'docno': "1",
            'description_doc': 'Report PM ' + engineer,
            'jobno': a_href.split('jobno=')[1].split('&')[0]
        })
        
        # Open file once and keep it in memory
        report_path = os.path.join(root_dir, 'REPORTS', report_name + '.pdf')
        with open(report_path, 'rb') as file:
            file_content = file.read()
            
        files = {'document_File': (report_name + '.pdf', file_content, 'application/pdf')}
        
        # Upload file
        post_url = f'https://nsmart.nhealth-asia.com/MTDPDB01/pm/maintain08.php?{a_href}&ccsForm=Maindocattache1'
        response = requests.post(
            post_url, headers=headers, cookies=cookies, data=form_data, files=files, verify=False, timeout=15
        )
        response.raise_for_status()
        response.encoding = "tis-620"
        
        # Check result
        soup = BeautifulSoup(response.text, "lxml")
        result_tables = soup.find_all('table', {'class': 'Header'})
        
        if len(result_tables) < 2:
            return {"status": 'fail', 'text': 'Fail to Attach PM file'}
        
        result_th = result_tables[1].find('th')
        if result_th.text.strip() != 'Total : 0 records':
            # Take screenshot if enabled
            if screenshot:
                get_screen_shot(soup, 'close_pm_css.css', result_th.text.strip(), "PM", code)
            
            return {"status": 'ok', 'text': 'Attach PM file : ' + result_th.text.strip()}
        else:
            return {"status": 'fail', 'text': 'Fail to Attach PM file'}
            
    except requests.exceptions.RequestException as e:
        logging.error(f"Network error in attachFilePM: {e}")
        time.sleep(3)  # Reduced sleep time
        return attachFilePM(file_name_list, row)
    except Exception as e:
        logging.error(f"Error in attachFilePM on line {sys.exc_info()[-1].tb_lineno}: {e}")
        return {"status": 'fail', 'text': f'Error: {str(e)}'}


def attachFileCAL(file_name_list, row):
    global SEARCH_KEY
    # Early returns for no-work, already-attached, or no-need cases
    if row['DATE-CAL'] == 'nan' or row['DATE-CAL'] == '':
        return {"status": 'ok', 'text': 'No CAL Work', 'nosave': True}
    if row['CAL-ATTACH-STATUS'].lower() == 'success':
        return {"status": 'done', 'text': row['CAL-ATTACH-STATUS'], 'nosave': True}
    if row['ATTACH-FILE-CAL'] != 'yes':
        return {"status": 'ok', 'text': 'No need to attach CAL file', 'nosave': True}
    
    # Check session and login if needed
    if cookies['PHPSESSID'] is None:
        set_login()
        return attachFileCAL(file_name_list, row)
    
    # Find report file
    code = row[SEARCH_KEY[0]]
    now_year = row['YEAR']
    findname = [ele for ele in file_name_list if row['ID CODE'] + '_' + now_year + '_cal' in ele]
    if not findname:
        return {"status": 'fail', 'text': 'CAL File For Attach not found'}
    report_name = findname[0]
    
    # Get job details
    start_date = formatDate(row['START-PLAN'])
    end_date = formatDate(row['END-PLAN'])
    
    # Build URL for request
    url = (f"https://nsmart.nhealth-asia.com/MTDPDB01/caliber/caliber03.php?s_byear={now_year}"
           f"&s_jobdate={start_date}&s_to_date={end_date}&{SEARCH_KEY[1]}={code}")
    
    try:
        # Use timeout to prevent hanging requests
        response = requests.get(
            url, headers=headers, cookies=cookies, verify=False, timeout=10
        )
        response.raise_for_status()
        response.encoding = "tis-620"
        
        soup = BeautifulSoup(response.text, "lxml")
        table = soup.find("table", {"class", "Grid"})
        
        # Handle authentication or empty results
        if not table:
            set_login()
            return attachFileCAL(file_name_list, row)
            
        tr = table.find('tr', {"class", "Row"})
        if not tr:
            return {"status": 'fail', 'text': 'CAL Work not found'}
        
        # Get attachment page
        a_href = tr.find('a')['href'].split('?')[1]
        attach_url = f"https://nsmart.nhealth-asia.com/MTDPDB01/caliber/caliber03_5.php?{a_href}&ccsForm=caliber_jobs_tech%3AEdit"
        response = requests.get(
            attach_url, headers=headers, cookies=cookies, verify=False, timeout=10
        )
        response.raise_for_status()
        response.encoding = "tis-620"
        
        soup = BeautifulSoup(response.text, "lxml")
        form = soup.find('form', {'name': 'Post'})
        
        if not form:
            set_login()
            return attachFileCAL(file_name_list, row)
        
        # Find and delete existing files
        file_tr = soup.find_all('table', {'class': 'Grid'})[1].find_all('tr', {'class': 'Row'})
        file_count = []
        engineer = row['ENGINEER'].upper()
        
        for tr in file_tr:
            if tr.find_all('td')[2].text.strip() == 'Report CAL ' + engineer:
                file_count.append(tr.find_all('td')[0].find('a')['href'])
        
        # Batch delete old files
        if file_count:
            for url, i in zip(file_count, range(len(file_count))):
                logging.info(f'Deleting old CAL file: {i+1}/{len(file_count)} {code}')
                try:
                    # Get delete form
                    del_url = f"https://nsmart.nhealth-asia.com/MTDPDB01/caliber/{url}"
                    del_response = requests.get(
                        del_url, headers=headers, cookies=cookies, verify=False, timeout=10
                    )
                    del_response.raise_for_status()
                    del_response.encoding = "tis-620"
                    
                    del_soup = BeautifulSoup(del_response.text, "lxml")
                    del_form = del_soup.find('form', {'name': 'Post'})
                    
                    if not del_form:
                        continue
                    
                    # Prepare delete form data
                    form_data = {}
                    for inp in del_form.findAll('input'):
                        try:
                            form_data[inp['name']] = inp['value']
                        except KeyError:
                            continue
                    
                    # Remove update button and add delete button coordinates
                    if "Button_Update" in form_data:
                        form_data.pop("Button_Update")
                    form_data['Button_Delete.x'] = str(random.randint(10, 50))
                    form_data['Button_Delete.y'] = str(random.randint(5, 20))
                    
                    # Submit delete request
                    del_post_url = f"https://nsmart.nhealth-asia.com/MTDPDB01/caliber/{url}&ccsForm=Caliberdocattache1%3AEdit"
                    del_response = requests.post(
                        del_post_url, headers=headers, cookies=cookies, data=form_data, verify=False, timeout=10
                    )
                    del_response.raise_for_status()
                    
                except requests.exceptions.RequestException as e:
                    logging.error(f"Error deleting CAL file: {e}")
                    time.sleep(2)  # Shorter sleep time
            
            print(f'[red]Delete old CAL file: {len(file_count)} file(s) success[/red] [yellow]{code}[/yellow]')
        
        # Prepare form data for file upload - simpler approach
        form_data = {
            'docno': "1",
            'description_doc': 'Report CAL ' + engineer,
            'jobno': a_href.split('jobno=')[1].split('&')[0]
        }
        
        # Open file once and read content directly
        report_path = os.path.join(root_dir, 'REPORTS', report_name + '.pdf')
        with open(report_path, 'rb') as file:
            file_content = file.read()
            
        files = {'document_File': ('report.pdf', file_content, 'application/pdf')}
        
        # Upload file
        post_url = f"https://nsmart.nhealth-asia.com/MTDPDB01/caliber/caliber03_5.php?{a_href}&ccsForm=Caliberdocattache1"
        response = requests.post(
            post_url, headers=headers, cookies=cookies, data=form_data, files=files, verify=False, timeout=15
        )
        response.raise_for_status()
        response.encoding = "tis-620"
        
        # Check result
        soup = BeautifulSoup(response.text, "lxml")
        result_tables = soup.find_all('table', {'class': 'Header'})
        
        if len(result_tables) < 2:
            return {"status": 'fail', 'text': 'Fail to Attach CAL file'}
        
        result_th = result_tables[1].find('th')
        if result_th.text.strip() != 'Total : 0 records':
            # Take screenshot if enabled
            if screenshot:
                get_screen_shot(soup, 'close_cal_css.css', result_th.text.strip(), "CAL", code)
            
            return {"status": 'ok', 'text': 'Attach CAL file : ' + result_th.text.strip()}
        else:
            return {"status": 'fail', 'text': 'Fail to Attach CAL file'}
            
    except requests.exceptions.RequestException as e:
        logging.error(f"Network error in attachFileCAL: {e}")
        time.sleep(3)  # Reduced sleep time
        return attachFileCAL(file_name_list, row)
    except Exception as e:
        logging.error(f"Error in attachFileCAL on line {sys.exc_info()[-1].tb_lineno}: {e}")
        return {"status": 'fail', 'text': f'Error: {str(e)}'}

screenshot = False


def read_file(option=None):
    global SEARCH_KEY
    # Create a stylish selection prompt
    console = Console()
    console.print(Panel(
        "[bold yellow]Select the search field to use:[/bold yellow]\n\n"
        "  [bold cyan]1[/bold cyan] [white]|[/white] [bold magenta]ID CODE[/bold magenta]    - Search by equipment identification code\n"
        "  [bold cyan]2[/bold cyan] [white]|[/white] [bold magenta]ITEM NO.[/bold magenta]   - Search by item number in database",
        title="[bold white]🔍 Search Configuration[/bold white]",
        border_style="cyan",
        expand=False,
        padding=(1, 2)
    ))
    
    # Add sparkles for fabulousness
    console.print("[bold yellow]✨ Enter your selection (1/2):[/bold yellow] ", end="")
    SEARCH_KEY = input()
    if SEARCH_KEY == '1':
        SEARCH_KEY = ['ID CODE', 's_sap_code']
    elif SEARCH_KEY == '2':
        SEARCH_KEY = ['ITEM NO', 's_code']
    
    # Read data and prepare environment
    df = read_excel_file()
    dir_path = os.path.dirname(sys.executable) if getattr(sys, 'frozen', False) else os.path.dirname(os.path.realpath(__file__))
    path = os.path.join(dir_path, 'REPORTS')
    
    # Get PDF filenames once and store in a set for faster lookups
    file_name_list = set()
    for file in os.listdir(path):
        if file.endswith('.pdf'):
            file_name_list.add(file.replace('.pdf', ''))
    
    try:
        print('[red]Reading file...[/red]')
        
        # Create backup before processing
        excel_path = os.path.join(excel_folder, excel_file_name).replace('//', '/')
        backup_path = os.path.join(excel_folder, 'backup' + excel_file_name).replace('//', '/')
        shutil.copy(excel_path, backup_path)
        
        # Pre-count statistics for display (only once, not in the loop)
        pass_pm = sum(1 for x in df['PM-CLOSED'] if str(x).lower() != 'success')
        pass_cal = sum(1 for x in df['CAL-CLOSED'] if str(x).lower() != 'success')
        attach_pm = sum(1 for x in df['PM-ATTACH-STATUS'] if str(x).lower() != 'success')
        attach_cal = sum(1 for x in df['CAL-ATTACH-STATUS'] if str(x).lower() != 'success')
        
        # Display statistics table
        table = Table(title=f'Total Medical Devices Found: {len(df)} devices', 
                     title_justify='center', title_style='bold magenta')
        table.add_column('#', justify='left', style='cyan', no_wrap=True)
        table.add_column('Number of Devices', justify='right', style='green', no_wrap=True)
        if option == 'close_pm_cal' or option is None:
            if pass_pm == 0 and pass_cal == 0:
                table.add_row('PM Jobs Closed', 'All Closed')
                table.add_row('CAL Jobs Closed', 'All Closed')
            else:
                table.add_row('PM Jobs Not Closed', str(pass_pm))
                table.add_row('CAL Jobs Not Closed', str(pass_cal))
        if option == 'attach_pm_cal' or option is None:
            if attach_pm == 0 and attach_cal == 0:
                table.add_row('PM Jobs Attached', 'All Attached')
                table.add_row('CAL Jobs Attached', 'All Attached')
            else:
                table.add_row('PM Jobs Not Attached', str(attach_pm))
                table.add_row('CAL Jobs Not Attached', str(attach_cal))
            
        Console().print(table)
        logging.info(f'Total Medical Devices Found: {len(df)} devices')
        # Early return if all jobs are already closed or attached based on selected option
        early_return = False
        
        if option == 'close_pm_cal' and pass_pm == 0 and pass_cal == 0:
            logging.info('All PM and CAL jobs are already closed')
            early_return = True
        elif option == 'attach_pm_cal' and attach_pm == 0 and attach_cal == 0:
            logging.info('All PM and CAL jobs are already attached')
            early_return = True
        elif option is None and pass_pm == 0 and pass_cal == 0 and attach_pm == 0 and attach_cal == 0:
            logging.info('All PM and CAL jobs are already closed and attached')
            early_return = True
            
        if early_return:
            print('Press any key to return to the main menu: ', end='')
            input()
            os.system('cls' if os.name == 'nt' else 'clear')
            init_text = pyfiglet.figlet_format("Start Process...")
            print(init_text)
            showmenu()
            return
        
        # Sparkly confirmation prompt
        console.print("[bold red]✨ Do you want to proceed? (Y/N):[/bold red] ", end="")
        if input().lower() != 'y':
            console.print("[bold red]Operation cancelled by user[/bold red]")
            time.sleep(1)
            os.system('cls' if os.name == 'nt' else 'clear')
            print(init_text)
            showmenu()
            return
        
        # Screenshot option with fancy prompt
        screenshot_panel = Panel(
            "[bold cyan]Screenshots can help with:[/bold cyan]\n\n"
            "  • [yellow]Documenting successful operations[/yellow]\n"
            "  • [yellow]Troubleshooting failed tasks[/yellow]\n"
            "  • [yellow]Providing evidence of completion[/yellow]",
            title="[bold white]📸 Screenshot Option[/bold white]",
            border_style="blue",
            expand=False,
            padding=(1, 2)
        )
        console.print(screenshot_panel)
        
        console.print("[bold blue]✨ Save screenshots of completed tasks? (Y/N):[/bold blue] ", end="")
        global screenshot
        screenshot = input().lower() == 'y'
        
        if screenshot:
            screenshot_dir = os.path.join(root_dir, 'SCREENSHOT')
            if not os.path.exists(screenshot_dir):
                os.makedirs(screenshot_dir)
                console.print(f"[green]Created screenshot directory at: [/green][blue]{screenshot_dir}[/blue]")
            else:
                console.print(f"[green]Screenshots will be saved to: [/green][blue]{screenshot_dir}[/blue]")
            
        # Countdown animation
        console.print("\n[bold magenta]Starting process in:[/bold magenta]")
        for i in range(3, 0, -1):
            console.print(f"[bold yellow]{i}...[/bold yellow]")
            time.sleep(0.5)
            
        # Start processing with fabulous header
        init_text = pyfiglet.figlet_format("Start Process...", font="slant")
        console.print(f"[bold rainbow]{init_text}[/bold rainbow]")
        print(init_text)
        start_time = time.time()
        
        # Convert file_name_list to list for compatibility with existing functions
        file_name_list = list(file_name_list)
        
        # Pre-calculate what needs to be processed
        to_process = []
        for idx, row in df.iterrows():
            # Clean the data once before processing
            row_data = row.copy()
            row_data[SEARCH_KEY[0]] = str(row_data[SEARCH_KEY[0]]).replace('.0', '')
            row_data['YEAR'] = str(row_data['YEAR']).replace('.0', '')
            
            # Skip rows with no ID code
            if row_data[SEARCH_KEY[0]] == 'nan':
                continue
                
            # Add row to processing list
            to_process.append((idx, row_data))
        
        # Set larger chunk size for better performance
        chunk_size = 20
        
        try:
            with Progress() as progress:
                task = progress.add_task("[red]Processing...[/red]", total=len(df))
                
                # Process in chunks with ThreadPoolExecutor
                for i in range(0, len(to_process), chunk_size):
                    chunk = to_process[i:min(i+chunk_size, len(to_process))]
                    
                    with ThreadPoolExecutor(max_workers=min(20, len(chunk) * 4)) as executor:
                        futures = []
                        
                        # Submit all tasks for this chunk
                        for j, (idx, row_data) in enumerate(chunk):
                            current_index = i + j + 1
                            print(f'[green]{current_index}[/green] / [blue]{len(to_process)}[/blue] [yellow]Processing[/yellow] [light blue]{row_data[SEARCH_KEY[0]]}[/light blue]')
                            logging.info(f'Processing: {row_data[SEARCH_KEY[0]]}')
                            
                            # Submit appropriate tasks based on option
                            pm_future = cal_future = pm_attach_future = cal_attach_future = None
                            
                            if option == 'close_pm_cal' or option is None:
                                pm_future = executor.submit(closePM, row_data)
                                cal_future = executor.submit(closeCAL, row_data)
                            
                            if option == 'attach_pm_cal' or option is None:
                                pm_attach_future = executor.submit(attachFilePM, file_name_list, row_data)
                                cal_attach_future = executor.submit(attachFileCAL, file_name_list, row_data)
                            
                            futures.append((idx, pm_future, cal_future, pm_attach_future, cal_attach_future, row_data[SEARCH_KEY[0]]))
                        
                        # Process results as they complete
                        for idx, pm_future, cal_future, pm_attach_future, cal_attach_future, id_code in futures:
                            issave = False
                            
                            # Process PM results
                            if pm_future:
                                process_result_pm = pm_future.result()
                                if process_result_pm.get('status') in ('ok', 'done'):
                                    issave = not (issave == False and process_result_pm.get('nosave') == True)
                                    result_text = "SUCCESS" if "Completed" in process_result_pm.get('text') else process_result_pm.get('text')
                                    df.at[idx, 'PM-CLOSED'] = result_text
                                    
                                    if process_result_pm.get('status') == 'ok':
                                        print(f'[green]{process_result_pm.get("text")}[/green] - [blue]{id_code}[/blue]')
                                        logging.info(f'PM-CLOSED: {process_result_pm.get("text")} - ID: {id_code}')
                                else:
                                    print(f'[red]{process_result_pm.get("text")}[/red] - [blue]{id_code}[/blue]')
                                    logging.error(f'PM-CLOSED: {process_result_pm.get("text")} - ID: {id_code}')
                                    df.at[idx, 'PM-CLOSED'] = process_result_pm.get('text')
                            
                            # Process CAL results
                            if cal_future:
                                process_result_cal = cal_future.result()
                                if process_result_cal.get('status') in ('ok', 'done'):
                                    issave = not (issave == False and process_result_cal.get('nosave') == True)
                                    result_text = "SUCCESS" if "Completed" in process_result_cal.get('text') else process_result_cal.get('text')
                                    df.at[idx, 'CAL-CLOSED'] = result_text
                                    
                                    if process_result_cal.get('status') == 'ok':
                                        print(f'[green]{process_result_cal.get("text")}[/green] - [blue]{id_code}[/blue]')
                                        logging.info(f'CAL-CLOSED: {process_result_cal.get("text")} - ID: {id_code}')
                                elif process_result_cal.get('status') != 'done':
                                    print(f'[red]{process_result_cal.get("text")}[/red] - [blue]{id_code}[/blue]')
                                    logging.error(f'CAL-CLOSED: {process_result_cal.get("text")} - ID: {id_code}')
                                    df.at[idx, 'CAL-CLOSED'] = process_result_cal.get('text')
                            
                            # Process PM attachment results
                            if pm_attach_future:
                                process_result_pm_attach = pm_attach_future.result()
                                if process_result_pm_attach.get('status') in ('ok', 'done'):
                                    issave = not (issave == False and process_result_pm_attach.get('nosave') == True)
                                    result_text = "SUCCESS" if process_result_pm_attach.get('text').startswith('Attach') else process_result_pm_attach.get('text')
                                    df.at[idx, 'PM-ATTACH-STATUS'] = result_text
                                    
                                    if process_result_pm_attach.get('status') == 'ok':
                                        print(f'[green]{process_result_pm_attach.get("text")}[/green] - [blue]{id_code}[/blue]')
                                        logging.info(f'PM-ATTACH-STATUS: {process_result_pm_attach.get("text")} - ID: {id_code}')
                                else:
                                    print(f'[red]{process_result_pm_attach.get("text")}[/red] - [blue]{id_code}[/blue]')
                                    logging.error(f'PM-ATTACH-STATUS: {process_result_pm_attach.get("text")} - ID: {id_code}')
                                    df.at[idx, 'PM-ATTACH-STATUS'] = process_result_pm_attach.get('text')
                            
                            # Process CAL attachment results
                            if cal_attach_future:
                                process_result_cal_attach = cal_attach_future.result()
                                if process_result_cal_attach.get('status') in ('ok', 'done'):
                                    issave = not (issave == False and process_result_cal_attach.get('nosave') == True)
                                    result_text = "SUCCESS" if process_result_cal_attach.get('text').startswith('Attach') else process_result_cal_attach.get('text')
                                    df.at[idx, 'CAL-ATTACH-STATUS'] = result_text
                                    
                                    if process_result_cal_attach.get('status') == 'ok':
                                        print(f'[green]{process_result_cal_attach.get("text")}[/green] - [blue]{id_code}[/blue]')
                                        logging.info(f'CAL-ATTACH-STATUS: {process_result_cal_attach.get("text")} - ID: {id_code}')
                                else:
                                    print(f'[red]{process_result_cal_attach.get("text")}[/red] - [blue]{id_code}[/blue]')
                                    logging.error(f'CAL-ATTACH-STATUS: {process_result_cal_attach.get("text")} - ID: {id_code}')
                                    df.at[idx, 'CAL-ATTACH-STATUS'] = process_result_cal_attach.get('text')
                            
                            # Handle special cases based on option
                            if option == 'close_pm_cal':
                                curr_pm = df.at[idx, 'PM-ATTACH-STATUS']
                                curr_cal = df.at[idx, 'CAL-ATTACH-STATUS']
                                df.at[idx, 'PM-ATTACH-STATUS'] = curr_pm if str(curr_pm).lower() != 'nan' else ''
                                df.at[idx, 'CAL-ATTACH-STATUS'] = curr_cal if str(curr_cal).lower() != 'nan' else ''
                            elif option == 'attach_pm_cal':
                                curr_pm = df.at[idx, 'PM-CLOSED']
                                curr_cal = df.at[idx, 'CAL-CLOSED']
                                df.at[idx, 'PM-CLOSED'] = curr_pm if str(curr_pm).lower() != 'nan' else ''
                                df.at[idx, 'CAL-CLOSED'] = curr_cal if str(curr_cal).lower() != 'nan' else ''
                            
                            progress.update(task, advance=1)
                    
                    # Save after each chunk to prevent data loss
                    print('[green]===========Saving Chunk Results=============[/green]')
                    logging.info('===========Saving Chunk Results=============')
                    with pd.ExcelWriter(excel_path, mode='w', engine='openpyxl') as writer:
                        # Replace 'nan' values with empty strings
                        df_save = df.replace('^nan$', '', regex=True)
                        # Save as strings to preserve formatting
                        df_save.to_excel(writer, sheet_name='Sheet1', index=False)
        
        except Exception as e:
            print(f"An error occurred on line {sys.exc_info()[-1].tb_lineno}: {e}")
            logging.error(f"An error occurred on line {sys.exc_info()[-1].tb_lineno}: {e}")
        
        finally:
            # Final save
            print('[green]=============Save Result to Excel==============[/green]')
            logging.info('=============Save Result to Excel==============')
            
            with pd.ExcelWriter(excel_path, mode='w', engine='openpyxl') as writer:
                # Replace 'nan' values with empty strings and format as text
                df_final = df.replace('^nan$', '', regex=True)
                df_final.to_excel(writer, sheet_name='Sheet1', index=False)
            
            # Show completion information
            end_time = time.time()
            elapsed_time = end_time - start_time
            print('\n[green]Closed jobs and attached files successfully[/green]')
            print(f'[green]Completed in[/green]: [yellow]{convertTime(elapsed_time)}[/yellow]')
            logging.info('Closed jobs and attached files successfully')
            logging.info(f'Completed in: {convertTime(elapsed_time)}')
            
            print('Press any button to return to the main menu: ', end='')
            input()
            
            # Return to main menu
            os.system('cls' if os.name == 'nt' else 'clear')
            print(init_text)
            showmenu()
    
    except Exception as e:
        print(f"An error occurred on line {sys.exc_info()[-1].tb_lineno}: {e}")
        logging.error(f"An error occurred on line {sys.exc_info()[-1].tb_lineno}: {e}")
        print('[red]Error[/red]')
        
        print('Press any button to return to the main menu: ', end='')
        input()
        
        os.system('cls' if os.name == 'nt' else 'clear')
        print(init_text)
        showmenu()

def convertDate(date):
    if date is None or len(date) == 0 or date == '-' or date[0] == '':
        return ""
    date = date[0].split(':')[1].strip()
    months_str = ['JANUARY', 'FEBRUARY', 'MARCH', 'APRIL', 'MAY', 'JUNE',
                  'JULY', 'AUGUST', 'SEPTEMBER', 'OCTOBER', 'NOVEMBER', 'DECEMBER']
    months_str_short = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun',
                        'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
    if date.find(' ') > -1:
        date = date.split(' ')
        if len(date) == 1:
            return ""
        if len(date) == 3:
            month = months_str.index(date[1].upper()) + 1
            return date[0] + '/' + str(month) + '/' + date[2]
    elif date.find('-') > -1:
        date = date.split('-')
        if len(date) == 1:
            return ""
        if len(date) == 3:
            month = months_str_short.index(date[1].capitalize()) + 1
            date[0] = '0' + date[0] if len(date[0]) == 1 else date[0]
            date[1] = '0' + date[1] if len(date[1]) == 1 else date[1]
            date[2] = '20' + date[2] if len(date[2]) == 2 else date[2]
            return date[0] + '/' + str(month) + '/' + date[2]

    return ' '.join(date)


def getYear(date1, date2):
    if date1 == '' and date2 == '':
        return ''
    date = ''
    if date1 == '':
        date = date2
    elif date2 == '':
        date = date1
    else:
        date = date1
    date = date.split('/')
    return date[2]


def getStartMonth(date1, date2):
    # convert date string to date object
    if date1 == '' and date2 == '':
        return ''
    date = ''
    if date1 == '':
        date = date2
    elif date2 == '':
        date = date1
    else:
        date = date1
    # get first day of month
    date = date.split('/')
    date[0] = '1'
    date_object = datetime.strptime('/'.join(date), '%d/%m/%Y')
    return date_object.strftime('%d/%m/%Y')


def getEndMonth(date1, date2):
    # convert date string to date object
    if date1 == '' and date2 == '':
        return ''
    date = ''
    if date1 == '':
        date = date2
    elif date2 == '':
        date = date1
    else:
        date = date1
    # get last day of month
    date = date.split('/')
    date[0] = str(calendar.monthrange(int(date[2]), int(date[1]))[1])
    date_object = datetime.strptime('/'.join(date), '%d/%m/%Y')
    return date_object.strftime('%d/%m/%Y')


def getTeamName(engineer):
    global emp_list
    engineer = engineer.lower()
    if emp_list is None or len(emp_list) == 0:
        load_empList()
    # find team name in emp_list
    for team in emp_list:
        if engineer in emp_list[team]:
            return emp_list[team]['option_name']
    return ''


def change_file_name():


    # Get the directory path
    dir_path = os.path.dirname(sys.executable) if getattr(sys, 'frozen', False) else os.path.dirname(os.path.realpath(__file__))
    report_path = os.path.join(dir_path, 'REPORTS')
    
    # Function to recursively move files from subfolders
    def move_file_to_root(folder):
        for file in os.listdir(folder):
            file_path = os.path.join(folder, file)
            if os.path.isdir(file_path):
                move_file_to_root(file_path)
            else:
                shutil.move(file_path, os.path.join(report_path, file))
        os.rmdir(folder)

    # Move files from subfolders to root directory
    for file in os.listdir(report_path):
        if os.path.isdir(os.path.join(report_path, file)):
            move_file_to_root(os.path.join(report_path, file))
    
    # Get list of files
    dir_list = os.listdir(report_path)
    name_arr = {}
    
    # Compile regex patterns once outside the loop
    id_code_pattern = re.compile(r'ID CODE.*\n.*$', re.MULTILINE)
    certificate_pattern = re.compile(r'Certificate', re.MULTILINE)
    cal_date_pattern = re.compile(r'CALIBRATED DATE.*\n.*$', re.MULTILINE)
    issue_date_pattern = re.compile(r'ISSUE DATE.*\n.*$', re.MULTILINE)
    pm_date_pattern = re.compile(r'PM. DATE.*\n.*$', re.MULTILINE)
    safety_pattern = re.compile(r'electricalsafetyanalyzer', re.MULTILINE)
    engineer_pattern = re.compile(r'Approved by.*\n.*$', re.MULTILINE)
    department_pattern = re.compile(r'LOCATION.*\n.*$', re.MULTILINE)
    
    # Process files in batches for better performance
    batch_size = 10
    total_files = len(dir_list)
    
    # Display initial information panel
    console = Console()
    info_panel = Panel(
        f"[cyan]Found [bold]{total_files}[/bold] files to process[/cyan]",
        title="[bold white]File Processing[/bold white]",
        border_style="cyan"
    )
    console.print(info_panel)
    
    # Track statistics for summary
    stats = {
        "cal_files": 0,
        "pm_files": 0,
        "skipped_files": 0,
        "error_files": 0    }    # Create a multi-part progress display
    with Progress(
        "[progress.description]{task.description}",
        "[progress.percentage]{task.percentage:>3.0f}%",
        BarColumn(),
        "[cyan]{task.completed}/{task.total}[/cyan]",
        "[progress.elapsed]Elapsed: {task.elapsed}",
        console=console,
        expand=True
    ) as progress:
        # Create the batch task first so it appears above the overall progress
        batch_task = progress.add_task("[bold yellow]Current Batch[/bold yellow]", total=batch_size, visible=False)
        # Add overall task last so it appears at the bottom
        overall_task = progress.add_task("[bold cyan]Overall Progress[/bold cyan]", total=total_files)
        
        # Process files in batches
        for batch_start in range(0, total_files, batch_size):
            batch_end = min(batch_start + batch_size, total_files)
            batch_files = dir_list[batch_start:batch_end]
            
            # Reset and update batch progress
            progress.update(batch_task, completed=0, total=len(batch_files), visible=True)
            progress.update(batch_task, description=f"[bold yellow]Processing Batch {batch_start//batch_size + 1}/{(total_files+batch_size-1)//batch_size}[/bold yellow]")
            
            for i, file_name in enumerate(batch_files):
                # Update file description
                current_index = batch_start + i + 1
                progress.update(overall_task, description=f"[bold cyan]File {current_index}/{total_files}: {file_name[:30]}{'...' if len(file_name) > 30 else ''}[/bold cyan]")
                
                source = os.path.join(report_path, file_name)
                
                # Skip non-PDF files
                if not file_name.endswith('.pdf'):
                    print(f'[red]Not a PDF file[/red] : [yellow]{file_name}[/yellow]')
                    progress.update(overall_task, advance=1)
                    progress.update(batch_task, advance=1)
                    stats["skipped_files"] += 1
                    continue
                
                try:
                    # Use context manager for file handling to ensure it's closed properly
                    with fitz.open(source) as file:
                        # Get text from first page only
                        page_text = file[0].get_text()
                        # Clean up text
                        page_text = re.sub(r'^-\n', '', page_text, flags=re.MULTILINE)
                        
                        # Extract equipment code
                        text = id_code_pattern.findall(page_text)
                        if not text:
                            print(f'[red]No equipment code found in PDF[/red] : [yellow]{file_name}[/yellow]')
                            progress.update(overall_task, advance=1)
                            progress.update(batch_task, advance=1)
                            stats["skipped_files"] += 1
                            continue
                            
                        code = text[0].split('\n')[1].replace(':', '').strip()
                        if '(' in code:
                            code = code.split('(')[0].strip()
                        
                        # Determine if this is calibration or PM
                        cal = certificate_pattern.findall(page_text)
                        year = None
                        
                        if cal:
                            # It's a calibration report
                            caldate = cal_date_pattern.findall(page_text)
                            issuedate = issue_date_pattern.findall(page_text)
                            year = getYear(convertDate(caldate), convertDate(issuedate))
                            
                            key = f"{code}#{year}"
                            if key not in name_arr:
                                name_arr[key] = {}
                                
                            name_arr[key]['cal'] = caldate
                            name_arr[key]['issue-cal'] = issuedate
                            new_name = f"{code}_{year}_cal.pdf"
                            stats["cal_files"] += 1
                        else:
                            # It's a PM report
                            pmdate = pm_date_pattern.findall(page_text)
                            issuedate = issue_date_pattern.findall(page_text)
                            year = getYear(convertDate(pmdate), convertDate(issuedate))
                            
                            key = f"{code}#{year}"
                            if key not in name_arr:
                                name_arr[key] = {}
                                
                            name_arr[key]['pm'] = pmdate
                            name_arr[key]['issue-pm'] = issuedate
                            new_name = f"{code}_{year}_pm.pdf"
                            stats["pm_files"] += 1
                        
                        # Check for safety information
                        lowercase_text = page_text.lower().replace('\n', '').replace(' ', '')
                        safety = safety_pattern.findall(lowercase_text)
                        name_arr[key]['safety'] = 'Electrical Safety Analyzer' if safety else '-'
                        
                        # Extract engineer information
                        engineer = engineer_pattern.findall(page_text)
                        if engineer:
                            engineer_name = engineer[0].replace('\n', '').split(':')[1].strip().replace('  ', ' ')
                            name_arr[key]['engineer'] = engineer_name
                        else:
                            name_arr[key]['engineer'] = '-'
                        
                        # Extract department information
                        department = department_pattern.findall(page_text)
                        if department:
                            dept_name = department[0].replace('\n', '').split(':')[1].strip().replace('  ', ' ')
                            name_arr[key]['department'] = dept_name
                        else:
                            name_arr[key]['department'] = '-'
                
                    # Rename the file with error handling
                    target_path = os.path.join(report_path, new_name)
                    try:
                        if new_name == file_name:
                            print(f'[grey42]{file_name}[/grey42] [yellow]>>>>[/yellow] [green]{new_name}[/green] [cyan](already named correctly)[/cyan]')
                        else:
                            if os.path.exists(target_path):
                                os.remove(target_path)
                            os.rename(source, target_path)
                            print(f'[grey42]{file_name}[/grey42] [yellow]>>>>[/yellow] [green]{new_name}[/green]')
                    except Exception as e:
                        print(f'[red]Error renaming file: {e}[/red]')
                        stats["error_files"] += 1
                
                except Exception as e:
                    print(f'[red]Error processing {file_name}: {e}[/red]')
                    stats["error_files"] += 1
                
                # Update progress
                progress.update(overall_task, advance=1)
                progress.update(batch_task, advance=1)
    
    # Display summary statistics
    summary_table = Table(title="File Processing Summary", show_header=True, box=box.ROUNDED)
    summary_table.add_column("Category", style="cyan")
    summary_table.add_column("Count", style="green", justify="right")
    summary_table.add_column("Percentage", style="yellow", justify="right")
    
    total_processed = stats["cal_files"] + stats["pm_files"]
    total_with_errors = total_processed + stats["skipped_files"] + stats["error_files"]
    
    summary_table.add_row("PM Files Processed", str(stats["pm_files"]), f"{stats['pm_files']/total_files*100:.1f}%")
    summary_table.add_row("CAL Files Processed", str(stats["cal_files"]), f"{stats['cal_files']/total_files*100:.1f}%")
    summary_table.add_row("Skipped Files", str(stats["skipped_files"]), f"{stats['skipped_files']/total_files*100:.1f}%")
    summary_table.add_row("Error Files", str(stats["error_files"]), f"{stats['error_files']/total_files*100:.1f}%")
    summary_table.add_row("Total Files", str(total_files), "100.0%")
    
    console.print(summary_table)
        
    # Prepare data for Excel
    unique_arr = []
    for i, key in enumerate(name_arr):
        value = name_arr[key]
        code, year = key.split('#')
        
        # Initialize with default values
        if 'cal' not in value:
            value['cal'] = ['']
        if 'pm' not in value:
            value['pm'] = ['']
        if 'engineer' not in value:
            value['engineer'] = '-'
        if 'department' not in value:
            value['department'] = '-'
        
        # Create row with proper fields
        tmp_arr = [''] * 26
        tmp_arr[0] = str(i+1)  # Item number
        tmp_arr[1] = ""        # Empty field
        tmp_arr[2] = code      # Equipment code
        tmp_arr[3] = getTeamName(value['engineer'])  # Team name
        tmp_arr[4] = value['engineer']  # Engineer name
        tmp_arr[5] = year      # Year
        
        # PM data
        pm_date = convertDate(value['pm'])
        tmp_arr[8] = pm_date
        
        if pm_date:
            tmp_arr[9] = convertDate(value['issue-pm']) or pm_date
            tmp_arr[14] = 'PM doable'
            tmp_arr[16] = 'pass'
            tmp_arr[19] = 'yes'
        
        # CAL data
        cal_date = convertDate(value['cal'])
        tmp_arr[10] = cal_date
        
        if cal_date:
            tmp_arr[11] = convertDate(value['issue-cal']) or cal_date
            tmp_arr[15] = 'Perform CAL'
            tmp_arr[17] = 'pass'
            tmp_arr[20] = 'yes'
        
        # Common fields
        tmp_arr[6] = getStartMonth(tmp_arr[8], tmp_arr[10])
        tmp_arr[7] = getEndMonth(tmp_arr[8], tmp_arr[10])
        tmp_arr[12] = confdata["SUP_ID"]
        tmp_arr[13] = confdata["SUP_NAME"]
        tmp_arr[18] = value.get('safety', '-')
        tmp_arr[25] = value.get('department', '-')
        
        unique_arr.append(tmp_arr)
    
    # Copy data to clipboard for Excel
    line_strings = ['\t'.join(line).replace('\n', '') for line in unique_arr]
    arr_string = '\r\n'.join(line_strings)
    pyperclip.copy(arr_string)
    
    print('\n[green]Successfully copied equipment codes[/green]')
    print(f'[cyan]Copied data for {len(unique_arr)} equipment items to clipboard[/cyan]')
    print('\n[green]File name change completed[/green]')
    print('\n[purple]Press any key to return to the main menu: [/purple]', end='')
    input()
    
    # Clear console and return to menu
    os.system('cls' if os.name == 'nt' else 'clear')
    print(init_text)
    showmenu()

def change_file_name_MICal():
    dir_path = ''
    if getattr(sys, 'frozen', False):
        dir_path = os.path.dirname(sys.executable)
    else:
        dir_path = os.path.dirname(os.path.realpath(__file__))
    report_path = os.path.join(dir_path, 'REPORTS')
    # check if path has subfolder
    subfolder = os.listdir(report_path)

    def move_file_to_root(folder):
        subfolder = os.listdir(folder)
        for file in subfolder:
            if os.path.isdir(os.path.join(folder, file)):
                move_file_to_root(os.path.join(folder, file))
            else:
                # move file to root folder
                shutil.move(os.path.join(folder, file),
                            os.path.join(report_path, file))
        os.rmdir(folder)

    if len(subfolder) > 0:
        # move file to root folder
        for file in subfolder:
            if os.path.isdir(os.path.join(report_path, file)):
                move_file_to_root(os.path.join(report_path, file))
    dir_list = os.listdir(report_path)
    name_arr = {}
    with Progress() as progress:
        task = progress.add_task(
            "[cyan]Renaming files[/cyan]", total=len(dir_list))
        for file_name in dir_list:
            source = os.path.join(report_path, file_name)
            # if file is not pdf
            if not file_name.endswith('.pdf'):
                print(
                    '[red]Not a PDF file[/red] : [yellow]{}[/yellow]'.format(file_name))
                progress.update(task, advance=1)
                continue

            # if file_name.find('_') == -1:
            file = fitz.open(source, filetype='pdf')
            page = file[0]
            # find text with regex /ID CODE.*\n.*$/gm
            page = page.get_text()
            # replace text in page with regex /^-\n/gm
            page = re.sub(r'^-\n', '', page, flags=re.MULTILINE)
            text = re.findall(r'ID No.\/Tag No.*\n.*$', page, re.MULTILINE)
            if len(text) == 0:
                print(
                    '[red]No equipment code found in PDF[/red] : [yellow]{}[/yellow]'.format(file_name))
                progress.update(task, advance=1)
                continue
            code = text[0].split('\n')[1].replace(':', '').strip()
            year = None
            caldate = re.findall(
                r'Date Calibrated*\n.*$', page, re.MULTILINE)
            issuedate = re.findall(
                r'Date Issued*\n.*$', page, re.MULTILINE)

            year = getYear(convertDate(caldate), convertDate(issuedate))
            if name_arr.get(code+'#'+year) is None:
                name_arr[code+'#'+year] = {}
            name_arr[code+'#'+year]['cal'] = caldate
            name_arr[code+'#'+year]['issue-cal'] = issuedate
            name = code + "_"+year
            name_arr[code+'#'+year]['pm'] = caldate
            name_arr[code+'#'+year]['issue-pm'] = issuedate
            name_arr[code+'#'+year]['safety'] = '-'
            name_arr[code + '#'+year]['engineer'] = '-'
            name_arr[code + '#'+year]['department'] = '-'
            file.close()
            try:
                os.rename(source, os.path.join(report_path, name+'_pm.pdf'))
            except Exception as e:
                # if already exist
                os.remove(os.path.join(report_path, name+'_pm.pdf'))
                os.rename(source, os.path.join(report_path, name+'_pm.pdf'))

            # copy file and rename to _cal.pdf
            shutil.copy(os.path.join(report_path, name+'_pm.pdf'),
                        os.path.join(report_path, name+'_cal.pdf'))

            # writer = PdfWriter(clone_from=os.path.join(report_path, name))
            # for page in writer.pages:
            #     page.compress_content_streams(level=9)
            # with open(os.path.join(report_path, name), 'wb') as f:
            #     writer.write(f)

            print('[grey42]{}[/grey42] [yellow]>>>>[/yellow] [green]{}[/green]'.format(
                file_name, name))
            progress.update(task, advance=1)
            # else:
            #     filename = file_name.split('_')
            #     if len(filename) > 1 and len(filename[1]) > 10:
            #         filename = "_".join(filename[1:])
            #         filename = re.sub(r'\s\(\d{1,}\)', '', filename)
            #         name = filename.split('(')[0]
            #         if (filename.split('(')[-1] != '1).pdf' and filename.split('(')[-1] != '2).pdf'):
            #             name_arr.append(name)
            #             if filename.split('(')[1].find('PM') > -1:
            #                 name = name + '_pm.pdf'
            #             else:
            #                 name = name + '_cal.pdf'
            #             print('เปลี่ยนชื่อไฟล์ [yellow]{}[/yellow] เป็น [yellow]{}[/yellow]'.format(
            #                 file_name, name))
            #             dest = os.path.join(path, name)
            #             os.rename(source, dest)

            # append name_arr to clipboard

            # convert name_arr to table
        # create array with length is 20 and fill with ''
        unique_arr = []
        for i, x in enumerate(name_arr):
            id = x
            value = name_arr[x]
            if value.get('cal') is None:
                value['cal'] = ['']
            if value.get('pm') is None:
                value['pm'] = ['']
            tmp_arr = [''] * 26
            tmp_arr[0] = str(i+1)
            tmp_arr[1] = ""
            tmp_arr[2] = id.split('#')[0]
            tmp_arr[3] = getTeamName(value['engineer'])
            tmp_arr[8] = convertDate(value['pm'])

            if tmp_arr[8] == '':
                tmp_arr[9] = ''
                tmp_arr[14] = ''
                tmp_arr[16] = ''
                tmp_arr[19] = ''
            else:
                tmp_arr[9] = convertDate(value['issue-pm'])
                if tmp_arr[9] == '' or tmp_arr[8] == '-':
                    tmp_arr[9] = convertDate(value['pm'])
                tmp_arr[14] = 'PM doable'
                tmp_arr[16] = 'pass'
                tmp_arr[19] = 'yes'
            tmp_arr[10] = convertDate(value['cal'])
            if tmp_arr[10] == '':
                tmp_arr[11] = ''
                tmp_arr[15] = ''
                tmp_arr[17] = ''
                tmp_arr[20] = ''
            else:
                tmp_arr[11] = convertDate(value['issue-cal'])
                if tmp_arr[11] == '' or tmp_arr[11] == '-':
                    tmp_arr[11] = convertDate(value['cal'])
                tmp_arr[15] = 'Perform CAL'
                tmp_arr[17] = 'pass'
                tmp_arr[20] = 'yes'
            tmp_arr[5] = id.split('#')[1]
            tmp_arr[6] = getStartMonth(tmp_arr[8], tmp_arr[10])
            tmp_arr[7] = getEndMonth(tmp_arr[8], tmp_arr[10])
            tmp_arr[12] = confdata["SUP_ID"]
            tmp_arr[13] = confdata["SUP_NAME"]
            tmp_arr[18] = value['safety']
            tmp_arr[4] = value['engineer']
            tmp_arr[25] = value['department']
            unique_arr.append(tmp_arr)

            # if value.get('cal') is not None and value.get('pm') is not None:
            #     cal = value.get('cal')[0].split('\n')[1].replace(':', '').strip()
            #     pm = value.get('pm')[0].split('\n')[1].replace(':', '').strip()
            #     name_arr[id] = {'cal': cal, 'pm': pm}
            # else:
            #     name_arr[id] = {'cal': 'nan', 'pm': 'nan'}
        # copy to cilpboard to paste in excel

        line_strings = []
        for line in unique_arr:
            line_strings.append('\t'.join(line).replace('\n', ''))
        arr_string = '\r\n'.join(line_strings)
        pyperclip.copy(arr_string)
        print('\n[green]Successfully copied equipment codes[/green]')
        print('\n[green]File name change completed[/green]')
        print('\n[purple]Press any key to return to the main menu: [/purple]', end='')
        input()
    # clear console
    os.system('cls' if os.name == 'nt' else 'clear')
    print(init_text)
    showmenu()


def re_init_app():
    global config
    global confdata
    global login_site
    global data
    config = open(os.path.join(root_dir, "CONFIG", "config.json"), "r")
    confdata = json.load(config)

    login_site = confdata["SITE"]
    print(login_site)
    data['user'] = confdata["USERNAME"]
    data['pass'] = confdata["PASSWORD"]
    cookies['PHPSESSID'] = None

    today = datetime.now()
    confdata['Last run'] = today.strftime('%d/%m/%Y')
    with open(os.path.join(root_dir, 'CONFIG', 'config.json'), 'w') as f:
        json.dump(confdata, f)
    # delete file calibarator_list.json
    for file in ['calibrator_list.json', 'tools_list.json', 'emp_list.json']:
        if os.path.exists(os.path.join(root_dir, 'CONFIG', file)):
            os.remove(os.path.join(
                root_dir, 'CONFIG', file))
    load_empList()
    load_calibrator_list()
    os.system('cls' if os.name == 'nt' else 'clear')
    print(init_text)

def check_default_browser():
    """Check the default browser on Windows"""
    try:
        # Query the registry for the default browser
        with winreg.OpenKey(winreg.HKEY_CURRENT_USER, r"Software\Microsoft\Windows\Shell\Associations\UrlAssociations\http\UserChoice") as key:
            prog_id = winreg.QueryValueEx(key, "ProgId")[0]
            
        # Get the browser name from ProgId
        browser_names = {
            'ChromeHTML': 'Google Chrome',
            'FirefoxURL': 'Mozilla Firefox',
            'MSEdgeHTM': 'Microsoft Edge',
            'IE.HTTP': 'Internet Explorer',
            'OperaStable': 'Opera',
            'BraveHTML': 'Brave Browser'
        }
        
        browser = browser_names.get(prog_id, f"Unknown ({prog_id})")
        print(f"🌐 Default Browser: {browser}")
        return browser
        
    except Exception as e:
        print(f"❌ Could not determine default browser: {e}")
        return None
    
def inject_javascript_if_domain_matches(driver, js_code, target_domain):
    current_url = driver.current_url
    parsed_url = urlparse(current_url)
    if parsed_url.netloc and parsed_url.netloc.endswith(target_domain) or parsed_url.netloc == target_domain:
        print(f"Injecting JS into: {current_url}")
        driver.execute_script(js_code)
    else:
        print(f"Skipping JS injection for: {current_url} (not in {target_domain})")

def showmenu():
    last_run_date = confdata.get('Last run')
    today = datetime.now()
    if (last_run_date != today.strftime('%d/%m/%Y')):
        re_init_app()
    # menu
    console = Console()
    
    # Create a centered, stylized header with borders
    # header = Panel(
    #     "[bold cyan]CES Assistant Menu[/bold cyan]",
    #     border_style="blue",
    #     expand=False,
    #     padding=(1, 10)
    # )
    # console.print(header)
    
    header2_text = (
        "[dim]• Session ID: " + cookies['PHPSESSID'][:8] + "...[/dim]\n"
        f"[dim]• User: {confdata['USERNAME']} | Site: {login_site}[/dim]\n"
        f"[dim]• Last run: {confdata.get('Last run', 'Never')}[/dim]"
    )
    
    header2 = Panel(
        header2_text,
        title="[bold white]Session Info[/bold white]",
        border_style="dim",
        expand=False,
        padding=(1, 1)
    )
    console.print(header2)
    # Create a two-column layout for the menu options
    main_table = Table(
        show_header=True, 
        header_style="bold yellow",
        box=box.ROUNDED,
        expand=True,
        show_lines=False,
        padding=(0, 1)
    )
    
    main_table.add_column("Option", justify="center", style="yellow", no_wrap=True)
    main_table.add_column("Description", style="cyan")
    
    # Add rows with improved descriptions
    main_table.add_row("[1]", "[bold magenta]Download ECERT Files[/bold magenta] (Copy script to clipboard)")
    main_table.add_row("[2]", "[bold green]Process & Rename Report Files[/bold green] (Extract data)")
    main_table.add_row("[3]", "[bold blue]Close PM & CAL Jobs[/bold blue] (Complete task records)")
    main_table.add_row("[4]", "[bold orange1]Attach PM & CAL Files[/bold orange1] (Link documents)")
    main_table.add_row("[5]", "[bold red]Complete Full Process[/bold red] (Close jobs + attach files)")
    main_table.add_row("[6]", "[bold cyan]Get Equipment Database[/bold cyan] (Export to Excel)")
    main_table.add_row("[7]", "[bold yellow]Reinitialize[/bold yellow] Application (Reset session)")
    
    # Add visual styles and information
    panel = Panel(
        main_table,
        title="[bold white]Available Functions[/bold white]",
        border_style="green",
        expand=False,
        padding=(1, 1)
    )
    console.print(panel)
    
    # Add helpful footer information
    
    # Input prompt with more visual indication
    console.print(
        "\n[bold purple]>>> Enter your selection (1-7):[/bold purple]", 
        end=" "
    )
    menu = input()
    if menu == '1':
        dir_path = ''
        if getattr(sys, 'frozen', False):
            dir_path = os.path.dirname(sys.executable)
        else:
            dir_path = os.path.dirname(os.path.realpath(__file__))
        # open file "download cert.txt" at SOURCE folder in notepad
        # os.system('notepad.exe "' + os.path.join(dir_path,
        #           'SOURCE', 'download cert.txt') + '"')
        # print('[green]เปิดสคริปต์สำหรับดาวน์โหลดไฟล์ ECERT[/green]')
        script_text = open(os.path.join(dir_path, 'SOURCE',
                                        'download cert.txt'), encoding='utf-8').read()
        sweetAlert2_text = requests.get('https://cdn.jsdelivr.net/npm/sweetalert2@11')
        if sweetAlert2_text.status_code == 200:
            script_text =sweetAlert2_text.text + script_text
        from selenium import webdriver
        from selenium.webdriver.chrome.service import Service as ChromeService
        driver = None
        default_browser = check_default_browser()
        if default_browser is None:
            console.print("[red]❌ Could not determine default browser. Please set it manually.[/red]")
            return
        if default_browser == 'Google Chrome':
            from webdriver_manager.chrome import ChromeDriverManager
            service = ChromeService(ChromeDriverManager().install())
            driver = webdriver.Chrome(service=service)
        elif default_browser == 'Mozilla Firefox':
            from webdriver_manager.firefox import GeckoDriverManager
            service = ChromeService(GeckoDriverManager().install())
            driver = webdriver.Firefox(service=service)
        elif default_browser == 'Microsoft Edge':
            from webdriver_manager.microsoft import EdgeChromiumDriverManager
            service = ChromeService(EdgeChromiumDriverManager().install())
            driver = webdriver.Edge(service=service)

        domain = 'necert.nhealth-asia.com/'
        
        driver.get(f'https://{domain}')
        try:
            WebDriverWait(driver, 1000).until(
                EC.presence_of_element_located((By.ID, 'plans_equipments_wrapper'))
            )  
        except TimeoutException:
            console.print("[red]❌ Login form not found. Please check the website URL.[/red]")
            driver.quit()
            return
        # Inject JavaScript to set cookies
        driver.execute_script(script_text)

        


        console.print(Panel(
            Align.center("[bold green]✓ Script executed in browser! ✓[/bold green]\n"
                 "[italic yellow]Download should start automatically. Check your downloads folder.[/italic yellow]"),
            title="[bold white]Download Tool Activated[/bold white]",
            border_style="green",
            padding=(1, 2)
        ))

        # Keep browser open until user decides to close
        console.print("\n[bold magenta]Press Enter to close the browser and return to main menu...[/bold magenta]")
        input()
        driver.quit()

        logging.info('ECERT download script executed successfully')
        os.system('cls' if os.name == 'nt' else 'clear')
        print(init_text)
        showmenu()
        # copy to clipboard
        # pyperclip.copy(script_text)
        # logging.info('User selected menu #1')
        # console.print(Panel(
        #     Align.center("[bold green]✓ Script copied to clipboard successfully! ✓[/bold green]\n"
        #          "[italic yellow]You can now paste it in the browser console on the ECERT page[/italic yellow]"),
        #     title="[bold white]Download Tool Ready[/bold white]",
        #     border_style="green",
        #     padding=(1, 2)
        # ))
        # logging.info('Script for downloading ECERT files copied successfully')
        # print('\nPress any key to return to the main menu: ', end='')
        # input()
        # os.system('cls' if os.name == 'nt' else 'clear')
        # print(init_text)
        # showmenu()
    elif menu == '2':
        logging.info('User selected menu #2')
        change_file_name()
        showmenu()
    elif menu == '6':
        logging.info('User selected menu #6')
        get_equipment_file()
        showmenu()

    elif menu == '7':
        logging.info('User selected menu #7')
        re_init_app()
        showmenu()
    else:
        load_empList()
        load_calibrator_list()
        load_tool_list('none')

        if menu == '3':
            logging.info('User selected menu #3')
            read_file('close_pm_cal')
        elif menu == '4':
            logging.info('User selected menu #4')
            read_file('attach_pm_cal')
        elif menu == '5':
            logging.info('User selected menu #5')
            read_file()

        else:
            print('Menu not found')
            showmenu()


try:
    showmenu()
    # change_file_name_MICal()
except Exception as e:
    print(f"An error occurred on line {sys.exc_info()[-1].tb_lineno}: {e}")
    logging.error(
        f"An error occurred on line {sys.exc_info()[-1].tb_lineno}: {e}")
